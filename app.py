"""
╔══════════════════════════════════════════════════════════════════════════╗
║   CO-PO Attainment & Academic Analytics System  — v3.0                  ║
║   NBA Accreditation Tool | All Fixes Applied                             ║
╠══════════════════════════════════════════════════════════════════════════╣
║  v3 Fixes:                                                               ║
║  • DI top%, bottom%, threshold — fully user-configurable                 ║
║  • Average rows: unmapped COs show "—" not 0.0                          ║
║  • DI column headers are dynamic (not hardcoded "27%")                   ║
║  • PDF: A4 portrait, reportlab, properly formatted pages                 ║
║  • Excel: correct averages, dynamic DI headers, avg row fix              ║
║  • UI: significantly improved layout and clarity                         ║
╚══════════════════════════════════════════════════════════════════════════╝
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from io import BytesIO
import math, textwrap, warnings
warnings.filterwarnings("ignore")

# reportlab for proper A4 PDF
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                    Spacer, PageBreak, Image, HRFlowable)
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

# ─── Page config ────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="CO-PO Attainment System | NBA v3",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ══════════════════════════════════════════════════════════════════════════
#  GLOBAL CSS
# ══════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800;900&family=JetBrains+Mono:wght@400;500&display=swap');

*{box-sizing:border-box}
html,body,[class*="css"]{font-family:'Inter',sans-serif!important}
.stApp{background:#f1f5fb}
.main .block-container{padding:1rem 1.6rem 2.4rem;max-width:100%}

/* Global card polish */
[data-testid="stVerticalBlock"] > [style*="flex-direction: column;"] > [data-testid="stVerticalBlock"]{
  border-radius:12px;
}

/* ── HEADER ── */
.nba-hdr{
  background:linear-gradient(135deg,#061231 0%,#0d2463 45%,#1740ad 100%);
  border-radius:18px;padding:26px 34px;margin-bottom:22px;
  display:flex;align-items:center;gap:22px;
  box-shadow:0 10px 40px rgba(6,18,49,.4);
  border:1px solid rgba(255,255,255,.07);
  position:relative;overflow:hidden;
}
.nba-hdr::before{
  content:'';position:absolute;right:-60px;top:-60px;
  width:260px;height:260px;
  border-radius:50%;
  background:radial-gradient(circle,rgba(255,255,255,.06) 0%,transparent 70%);
}
.hdr-icon{font-size:52px;line-height:1;z-index:1}
.hdr-text{z-index:1}
.hdr-title{color:#fff;font-size:22px;font-weight:900;margin:0;letter-spacing:-.5px}
.hdr-sub{color:#93b4f5;font-size:12px;margin:5px 0 0;font-weight:500;letter-spacing:.2px}
.hdr-pills{margin-left:auto;display:flex;gap:8px;flex-wrap:wrap;z-index:1}
.hdr-pill{
  background:rgba(255,255,255,.1);border:1px solid rgba(255,255,255,.18);
  color:#fff;padding:5px 14px;border-radius:20px;
  font-size:11px;font-weight:700;letter-spacing:.4px;white-space:nowrap;
}

/* ── SECTION TITLE ── */
.stitle{
  font-size:14px;font-weight:800;color:#061231;
  border-left:4px solid #1740ad;padding-left:11px;
  margin:18px 0 10px;letter-spacing:-.15px;
  display:flex;align-items:center;gap:6px;
}

/* ── METRIC STRIP ── */
.mstrip{display:grid;grid-template-columns:repeat(auto-fill,minmax(160px,1fr));gap:12px;margin-bottom:22px}
.mbox{
  background:linear-gradient(180deg,#ffffff 0%,#f8fbff 100%);
  border-radius:13px;padding:15px 16px;
  box-shadow:0 1px 4px rgba(6,18,49,.06),0 6px 20px rgba(6,18,49,.06);
  border:1px solid #e6ecfa;border-left:4px solid var(--ac,#1740ad);
  transition:transform .2s,box-shadow .2s;
}
.mbox:hover{transform:translateY(-2px);box-shadow:0 4px 20px rgba(0,0,0,.1)}
.ml{font-size:10px;font-weight:700;color:#6b7280;text-transform:uppercase;letter-spacing:.7px}
.mv{font-size:28px;font-weight:900;color:#061231;margin:5px 0 3px;line-height:1}
.ms{font-size:11.5px;color:#6b7280}

/* ── ALERTS ── */
.awarn{background:#fffbeb;border-left:4px solid #f59e0b;padding:12px 16px;
  border-radius:9px;margin:7px 0;font-size:13px;line-height:1.65;
  box-shadow:0 1px 4px rgba(245,158,11,.15)}
.agood{background:#f0fdf4;border-left:4px solid #22c55e;padding:12px 16px;
  border-radius:9px;margin:7px 0;font-size:13px;line-height:1.65;
  box-shadow:0 1px 4px rgba(34,197,94,.15)}
.ainfo{background:#eff6ff;border-left:4px solid #3b82f6;padding:12px 16px;
  border-radius:9px;margin:7px 0;font-size:13px;line-height:1.65;
  box-shadow:0 1px 4px rgba(59,130,246,.15)}
.adanger{background:#fff1f2;border-left:4px solid #f43f5e;padding:12px 16px;
  border-radius:9px;margin:7px 0;font-size:13px;line-height:1.65}

/* ── TABLES ── */
.nba-wrap{overflow-x:auto;border-radius:12px;
  box-shadow:0 2px 12px rgba(6,18,49,.08);margin:8px 0;border:1px solid #e6ecfa;background:#fff}
.nba-tbl{width:100%;border-collapse:collapse;font-size:12.5px}
.nba-tbl th{
  background:#061231;color:#fff;padding:9px 11px;
  text-align:center;font-weight:700;letter-spacing:.25px;
  white-space:nowrap;border-right:1px solid rgba(255,255,255,.09)}
.nba-tbl th.proc{background:#0d2463;text-align:left;min-width:140px}
.nba-tbl td{
  padding:8px 12px;text-align:center;border-bottom:1px solid #e5eaf5;
  border-right:1px solid #e5eaf5;font-variant-numeric:tabular-nums;
  font-size:12.5px}
.nba-tbl tr:nth-child(even) td{background:#f7f9fd}
.nba-tbl tr:hover td{background:#eef3ff;transition:background .15s}
.nba-tbl td.proc{text-align:left;font-weight:700;background:#f8fafc!important;
  white-space:nowrap;color:#061231}
.nba-tbl tr.avg-r td{
  background:#dde8ff!important;font-weight:800;color:#061231;
  border-top:2px solid #1740ad}
.nba-tbl tr.avg-r td.proc{background:#0d2463!important;color:#fff}
.zero{color:#c8c8c8;font-style:italic}
.hi{color:#1740ad;font-weight:800}
.warn{color:#dc2626;font-weight:700}
.good-di{color:#16a34a;font-weight:700}
.dash-cell{color:#c8c8c8}

/* ── LEVEL BADGES ── */
.lv0{background:#fee2e2;color:#b91c1c;border:1px solid #fca5a5;
  padding:1px 8px;border-radius:20px;font-size:11px;font-weight:800}
.lv1{background:#fef3c7;color:#92400e;border:1px solid #fcd34d;
  padding:1px 8px;border-radius:20px;font-size:11px;font-weight:800}
.lv2{background:#d1fae5;color:#065f46;border:1px solid #6ee7b7;
  padding:1px 8px;border-radius:20px;font-size:11px;font-weight:800}
.lv3{background:#dbeafe;color:#1e3a8a;border:1px solid #93c5fd;
  padding:1px 8px;border-radius:20px;font-size:11px;font-weight:800}

/* ── TABS ── */
.stTabs [data-baseweb="tab-list"]{
  background:linear-gradient(180deg,#ffffff 0%,#f8fbff 100%);
  border-radius:13px;padding:5px;gap:4px;border:1px solid #e6ecfa;
  box-shadow:0 2px 10px rgba(6,18,49,.08);margin-bottom:16px}
.stTabs [data-baseweb="tab"]{
  font-size:12.5px;font-weight:700;color:#4b5563;
  padding:10px 22px;border-radius:9px;transition:all .2s}
.stTabs [aria-selected="true"]{
  background:linear-gradient(135deg,#061231,#1740ad)!important;
  color:white!important;box-shadow:0 3px 10px rgba(23,64,173,.4)!important}

/* ── BUTTONS ── */
.stButton>button{
  background:linear-gradient(135deg,#061231,#1740ad)!important;
  color:white!important;border:none!important;border-radius:10px!important;
  font-weight:700!important;padding:10px 28px!important;font-size:13.5px!important;
  box-shadow:0 4px 16px rgba(23,64,173,.4)!important;transition:all .25s!important}
.stButton>button:hover{
  background:linear-gradient(135deg,#0d2463,#2554d4)!important;
  transform:translateY(-2px)!important;
  box-shadow:0 7px 22px rgba(23,64,173,.55)!important}
[data-testid="stDownloadButton"]>button{
  background:linear-gradient(135deg,#064e3b,#059669)!important;
  color:white!important;border:none!important;border-radius:10px!important;
  font-weight:700!important;font-size:13.5px!important;
  box-shadow:0 4px 14px rgba(5,150,105,.35)!important}
[data-testid="stDownloadButton"]>button:hover{
  background:linear-gradient(135deg,#065f46,#10b981)!important;
  transform:translateY(-2px)!important}

/* ── INPUTS ── */
.stNumberInput input,.stTextInput input{
  border-radius:9px!important;border:1.5px solid #dce5f5!important;
  font-size:12.5px!important;background:#fff!important}
.stNumberInput input:focus,.stTextInput input:focus{
  border-color:#1740ad!important;
  box-shadow:0 0 0 3px rgba(23,64,173,.12)!important}
.stTextArea textarea{border-radius:9px!important;border:1.5px solid #dce5f5!important}

/* Dataframe / editor polish */
[data-testid="stDataFrame"]{
  border:1px solid #e6ecfa;
  border-radius:10px;
  overflow:hidden;
  box-shadow:0 1px 8px rgba(6,18,49,.06);
}
[data-testid="stFileUploader"]{
  border:1.5px dashed #c7d4ef;
  border-radius:10px;
  padding:8px;
  background:rgba(255,255,255,.5);
}

/* ── SUBJECT BOX ── */
.subj-box{
  background:linear-gradient(135deg,#061231 0%,#0d2463 100%);
  border-radius:14px;padding:18px 20px;color:#fff;margin-bottom:14px;
  box-shadow:0 4px 20px rgba(6,18,49,.25);
}
.subj-box .sc{font-size:10px;font-weight:700;color:#93b4f5;
  text-transform:uppercase;letter-spacing:.9px;margin-bottom:4px}
.subj-box .sn{font-size:19px;font-weight:900;margin:0 0 12px}
.subj-box .co-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(280px,1fr));gap:6px;margin-top:8px}
.subj-box .co-item{
  background:rgba(255,255,255,.1);border:1px solid rgba(255,255,255,.15);
  padding:6px 12px;border-radius:8px;font-size:11.5px;line-height:1.4}
.subj-box .co-lbl{font-weight:700;color:#93b4f5}

/* ── DI CONFIG BOX ── */
.di-config{
  background:linear-gradient(135deg,#fefce8,#fef9c3);
  border:1px solid #fcd34d;border-radius:12px;padding:16px 20px;margin:12px 0}

/* ── INSIGHT CARD ── */
.insight-card{
  background:#fff;border-radius:12px;padding:18px 20px;
  box-shadow:0 2px 10px rgba(0,0,0,.06);border:1px solid #e5eaf5;
  margin:10px 0}

/* ── PROGRESS BAR ── */
.att-bar-wrap{display:flex;align-items:center;gap:10px;margin:3px 0}
.att-bar{height:8px;border-radius:4px;background:#e5eaf5;flex:1;overflow:hidden}
.att-bar-fill{height:100%;border-radius:4px;transition:width .4s}

/* ── SECTION DIVIDER ── */
.sec-div{border:none;border-top:2px solid #e5eaf5;margin:24px 0}

#MainMenu,footer,header{visibility:hidden}
::-webkit-scrollbar{width:6px;height:6px}
::-webkit-scrollbar-track{background:#f1f5fb}
::-webkit-scrollbar-thumb{background:#c0cbdf;border-radius:3px}
::-webkit-scrollbar-thumb:hover{background:#a0aec0}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════
#  CONSTANTS
# ══════════════════════════════════════════════════════════════════════════
NUM_COS  = 6
NUM_POS  = 12
NUM_PSOS = 3

INTERNAL_EXAMS = ["Unit Test 1", "Unit Test 2", "Unit Test 3", "Prelim"]
EXTERNAL_EXAMS = ["Insem", "Endsem"]
ALL_EXAMS      = INTERNAL_EXAMS + EXTERNAL_EXAMS
PRACTICAL_INTERNAL_EXAM = "Internal"
PRACTICAL_EXTERNAL_EXAM = "External Combined"

DEFAULT_MAX = {
    "Unit Test 1": 30,
    "Unit Test 2": 30,
    "Unit Test 3": 30,
    "Prelim":      70,
    "Insem":       30,
    "Endsem":      70,
}
DEFAULT_INTERNAL_WEIGHT   = 25.0
DEFAULT_UNIVERSITY_WEIGHT = 75.0
DEFAULT_DI_TOP_PCT        = 27.0
DEFAULT_DI_BOT_PCT        = 27.0
DEFAULT_DI_THRESHOLD      = 0.20
DEFAULT_MODE              = "Theory"
DEFAULT_PRACTICAL_INTERNAL_MAX = 25
DEFAULT_PRACTICAL_COMPONENTS = [
    {"name": "Practical", "max_marks": 50},
    {"name": "Term Work", "max_marks": 15},
    {"name": "Oral", "max_marks": 10},
]

DEFAULT_SUBJ_CODE = "304184"
DEFAULT_SUBJ_NAME = "Microcontrollers"
DEFAULT_CO_STMTS  = [
    "Understand the fundamentals of microcontroller and programming",
    "Interface various electronic components with microcontrollers",
    "Analyze the features of PIC 18F XXXX",
    "Describe the programming details in peripheral support",
    "Develop interfacing models according to applications",
    "Evaluate the serial communication details and interfaces",
]

DEFAULT_COPO = [
    [3, 1, 2, 1, 1, 0, 0, 0, 0, 0, 0, 1],
    [3, 0, 3, 1, 1, 0, 0, 0, 0, 0, 0, 1],
    [3, 3, 1, 1, 1, 0, 1, 0, 0, 0, 0, 1],
    [3, 0, 3, 1, 1, 0, 0, 0, 0, 0, 0, 1],
    [3, 0, 3, 1, 1, 1, 0, 0, 0, 0, 0, 1],
    [3, 1, 3, 1, 1, 1, 0, 0, 0, 0, 0, 1],
]

DEFAULT_COPSO = [
    [2, 1, 0],
    [2, 1, 0],
    [1, 2, 0],
    [1, 2, 1],
    [1, 1, 2],
    [1, 0, 2],
]

DEFAULT_COEXAM = [
    [1, 0, 0, 0, 1, 0],
    [1, 0, 0, 0, 1, 0],
    [0, 1, 0, 1, 0, 1],
    [0, 1, 0, 1, 0, 1],
    [0, 0, 1, 0, 0, 1],
    [0, 0, 1, 0, 0, 1],
]

# Colour constants
NAV = "0A1628"
BLU = "1740AD"
LBL = "EEF3FF"
EVN = "F7F9FD"
DI_LABEL = "DI (H-L)/Max Marks"


# ══════════════════════════════════════════════════════════════════════════
#  CALCULATION FUNCTIONS
# ══════════════════════════════════════════════════════════════════════════

def attainment_level(pct: float) -> int:
    if   pct < 40:  return 0
    elif pct <= 55: return 1
    elif pct <= 70: return 2
    else:           return 3


def exam_attainment(marks: np.ndarray, max_marks: float):
    """NBA formula: att = (0·P + 1·Q + 2·R + 3·S) / N"""
    marks  = np.asarray(marks, dtype=float)
    pcts   = (marks / max_marks) * 100.0
    levels = np.array([attainment_level(p) for p in pcts])
    n      = len(levels)
    if n == 0:
        return 0.0, levels
    P, Q, R, S = [(levels == k).sum() for k in range(4)]
    return float(0*P + 1*Q + 2*R + 3*S) / n, levels


def discrimination_index(marks: np.ndarray, max_marks: float, top_pct: float, bot_pct: float):
    """
    DI = (H_mean − L_mean) / Max Marks
    top_pct, bot_pct in 0–100 (e.g. 27 means top/bottom 27%)
    """
    arr = np.sort(np.asarray(marks, dtype=float))
    n   = len(arr)
    if n == 0:
        return 0.0, 0.0, 0.0, 0, 0
    max_marks = float(max_marks)
    k_top = max(1, int(np.ceil(n * top_pct / 100.0)))
    k_bot = max(1, int(np.ceil(n * bot_pct / 100.0)))
    H  = arr[-k_top:].mean()
    L  = arr[:k_bot].mean()
    DI = (H - L) / max_marks if max_marks > 0 else 0.0
    DI = min(max(float(DI), 0.0), 1.0)
    return round(float(DI), 4), round(float(H), 4), round(float(L), 4), int(k_top), int(k_bot)


def compute_co_attainments(exam_att_dict, coexam_mat, internal_exams, external_exams, all_exams):
    """Compute per-CO internal and external attainments from exam-level data."""
    co_int_att = np.zeros(NUM_COS)
    co_ext_att = np.zeros(NUM_COS)
    for i in range(NUM_COS):
        # Internal
        int_vals = []
        for exam in internal_exams:
            idx = all_exams.index(exam)
            if coexam_mat[i, idx] > 0:
                int_vals.append(exam_att_dict[exam])
        co_int_att[i] = np.mean(int_vals) if int_vals else np.nan

        # External
        ext_vals = []
        for exam in external_exams:
            idx = all_exams.index(exam)
            if coexam_mat[i, idx] > 0:
                ext_vals.append(exam_att_dict[exam])
        co_ext_att[i] = np.mean(ext_vals) if ext_vals else np.nan

    return co_int_att, co_ext_att


def compute_co_finals(co_int_att, co_ext_att, int_ratio, ext_ratio):
    """
    Final CO = int_ratio * Internal + ext_ratio * External
    If a CO has only internal or only external, use what's available (skip NaN side).
    """
    co_finals = np.zeros(NUM_COS)
    for i in range(NUM_COS):
        has_int = not np.isnan(co_int_att[i])
        has_ext = not np.isnan(co_ext_att[i])
        if has_int and has_ext:
            co_finals[i] = int_ratio * co_int_att[i] + ext_ratio * co_ext_att[i]
        elif has_int:
            co_finals[i] = co_int_att[i]
        elif has_ext:
            co_finals[i] = co_ext_att[i]
        else:
            co_finals[i] = 0.0
    return co_finals


def compute_po_pso(co_finals, mapping):
    """
    contrib[i,j] = (co_finals[i] / 3) * mapping[i,j]
    final[j] = mean(contrib[:,j]) over all COs
    """
    num_out = mapping.shape[1]
    contrib = np.zeros((NUM_COS, num_out))
    for i in range(NUM_COS):
        for j in range(num_out):
            contrib[i, j] = (co_finals[i] / 3.0) * mapping[i, j]
    return contrib, contrib.mean(axis=0)


def normalize_weights(iw, uw):
    total = float(iw) + float(uw)
    if total <= 0: return 0.25, 0.75
    return float(iw) / total, float(uw) / total


def assessment_labels(mode: str):
    if mode == "Practical":
        return "Internal", "External"
    return "Internal", "University"


def practical_component_col(idx: int) -> str:
    return f"External_Component_{idx+1}"


def practical_component_columns(components):
    return [practical_component_col(i) for i in range(len(components))]


def get_exam_sets(mode: str):
    if mode == "Practical":
        i_exams = [PRACTICAL_INTERNAL_EXAM]
        e_exams = [PRACTICAL_EXTERNAL_EXAM]
    else:
        i_exams = list(INTERNAL_EXAMS)
        e_exams = list(EXTERNAL_EXAMS)
    return i_exams, e_exams, i_exams + e_exams


def build_practical_calc_df(df_raw: pd.DataFrame, component_cols):
    df = df_raw.copy()
    for col in [PRACTICAL_INTERNAL_EXAM] + component_cols:
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)
    ext_sum = df[component_cols].sum(axis=1) if component_cols else 0
    out = pd.DataFrame({
        "Student Name": df.get("Student Name", pd.Series([f"Std_{i+1}" for i in range(len(df))])),
        PRACTICAL_INTERNAL_EXAM: df[PRACTICAL_INTERNAL_EXAM],
        PRACTICAL_EXTERNAL_EXAM: ext_sum,
    })
    return out


# ══════════════════════════════════════════════════════════════════════════
#  TABLE DATA-FRAME BUILDERS
# ══════════════════════════════════════════════════════════════════════════

def _safe(v, decimals=4):
    """Format float or return '—' for NaN/missing."""
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "—"
    return f"{float(v):.{decimals}f}"


def build_internal_df(co_names, coexam_mat, exam_att_dict, co_int_att, internal_exams, all_exams):
    rows = []
    for exam in internal_exams:
        idx = all_exams.index(exam)
        row = {"Process": exam}
        for i, co in enumerate(co_names):
            row[co] = round(float(exam_att_dict[exam]), 4) if coexam_mat[i, idx] > 0 else "—"
        rows.append(row)
    avg = {"Process": "Average Attainment of CO"}
    for i, co in enumerate(co_names):
        avg[co] = "—" if np.isnan(co_int_att[i]) else round(float(co_int_att[i]), 4)
    rows.append(avg)
    return pd.DataFrame(rows)


def build_external_df(co_names, coexam_mat, exam_att_dict, co_ext_att, external_exams, all_exams):
    rows = []
    for exam in external_exams:
        idx = all_exams.index(exam)
        row = {"Process": exam}
        for i, co in enumerate(co_names):
            row[co] = round(float(exam_att_dict[exam]), 4) if coexam_mat[i, idx] > 0 else "—"
        rows.append(row)
    avg = {"Process": "Average Attainment of CO"}
    for i, co in enumerate(co_names):
        avg[co] = "—" if np.isnan(co_ext_att[i]) else round(float(co_ext_att[i]), 4)
    rows.append(avg)
    return pd.DataFrame(rows)


def build_final_df(co_names, co_int_att, co_ext_att, co_finals, iw, uw, external_label="University"):
    ir, ur = normalize_weights(iw, uw)
    rows = []
    row1 = {"Process": f"{uw:.1f}% of {external_label}"}
    for i, co in enumerate(co_names):
        if np.isnan(co_ext_att[i]):
            row1[co] = "—"
        else:
            row1[co] = round(float(co_ext_att[i]) * ur, 4)

    row2 = {"Process": f"{iw:.1f}% of Internal"}
    for i, co in enumerate(co_names):
        if np.isnan(co_int_att[i]):
            row2[co] = "—"
        else:
            row2[co] = round(float(co_int_att[i]) * ir, 4)

    row3 = {"Process": "CO Attainment"}
    for i, co in enumerate(co_names):
        row3[co] = round(float(co_finals[i]), 4)

    return pd.DataFrame([row1, row2, row3])


def build_di_df(exam_di_dict, top_pct, bot_pct, threshold, internal_exams):
    rows = []
    for exam, d in exam_di_dict.items():
        rows.append({
            "Exam":                               exam,
            "Type":                               "Internal" if exam in internal_exams else "External",
            "N (Students)":                       d["N"],
            f"Top {top_pct:.0f}% Mean (H)":       round(float(d["H"]), 4),
            f"Bottom {bot_pct:.0f}% Mean (L)":    round(float(d["L"]), 4),
            f"k (Top)":                           d["k_top"],
            f"k (Bot)":                           d["k_bot"],
            DI_LABEL:                             round(float(d["DI"]), 4),
            "Status":                             "✓ Good" if d["DI"] >= threshold else "⚠ Low",
        })
    return pd.DataFrame(rows)


def build_po_df(co_names, co_finals, po_contrib, po_finals):
    po_names = [f"PO{j+1}" for j in range(NUM_POS)]
    rows = []
    for i, co in enumerate(co_names):
        row = {"CO": co, "CO Attainment": round(float(co_finals[i]), 4)}
        for j, po in enumerate(po_names):
            row[po] = round(float(po_contrib[i, j]), 4)
        rows.append(row)
    avg = {"CO": "Average", "CO Attainment": ""}
    for j, po in enumerate(po_names):
        avg[po] = round(float(po_finals[j]), 4)
    rows.append(avg)
    return pd.DataFrame(rows)


def build_pso_df(co_names, co_finals, pso_contrib, pso_finals):
    pso_names = [f"PSO{j+1}" for j in range(NUM_PSOS)]
    rows = []
    for i, co in enumerate(co_names):
        row = {"CO": co, "CO Attainment": round(float(co_finals[i]), 4)}
        for j, pso in enumerate(pso_names):
            row[pso] = round(float(pso_contrib[i, j]), 4)
        rows.append(row)
    avg = {"CO": "Average", "CO Attainment": ""}
    for j, pso in enumerate(pso_names):
        avg[pso] = round(float(pso_finals[j]), 4)
    rows.append(avg)
    return pd.DataFrame(rows)


# ══════════════════════════════════════════════════════════════════════════
#  HTML TABLE RENDERERS
# ══════════════════════════════════════════════════════════════════════════

def _v(val, decimals=4):
    """Format cell value for HTML table."""
    if isinstance(val, str):
        if val == "—":
            return '<span class="dash-cell">—</span>'
        return val
    if isinstance(val, float) and np.isnan(val):
        return '<span class="dash-cell">—</span>'
    try:
        f = float(val)
        if f == 0: return '<span class="zero">0</span>'
        return f"{f:.{decimals}f}"
    except:
        return str(val)


def _tbl(headers, rows_data, avg_row_data=None, proc_col=True):
    """Build an HTML NBA-style table."""
    thead_cells = "".join(
        f'<th class="proc">{h}</th>' if (i == 0 and proc_col) else f"<th>{h}</th>"
        for i, h in enumerate(headers)
    )
    tbody = ""
    for row in rows_data:
        cells = "".join(
            f'<td class="proc">{_v(v)}</td>' if (i == 0 and proc_col) else f"<td>{_v(v)}</td>"
            for i, v in enumerate(row)
        )
        tbody += f"<tr>{cells}</tr>"
    if avg_row_data:
        cells = "".join(
            f'<td class="proc">{_v(v)}</td>' if (i == 0 and proc_col) else f"<td>{_v(v)}</td>"
            for i, v in enumerate(avg_row_data)
        )
        tbody += f'<tr class="avg-r">{cells}</tr>'
    return f'<div class="nba-wrap"><table class="nba-tbl"><thead><tr>{thead_cells}</tr></thead><tbody>{tbody}</tbody></table></div>'


def render_internal_table(co_names, coexam_mat, exam_att_dict, co_int_att, internal_exams, all_exams):
    headers = ["Process"] + co_names
    rows = []
    for exam in internal_exams:
        idx = all_exams.index(exam)
        row = [exam]
        for i in range(NUM_COS):
            row.append(round(exam_att_dict[exam], 4) if coexam_mat[i, idx] > 0 else "—")
        rows.append(row)
    avg = ["Avg Attainment of CO"]
    for i in range(NUM_COS):
        avg.append("—" if np.isnan(co_int_att[i]) else round(float(co_int_att[i]), 4))
    return _tbl(headers, rows, avg)


def render_external_table(co_names, coexam_mat, exam_att_dict, co_ext_att, external_exams, all_exams):
    headers = ["Process"] + co_names
    rows = []
    for exam in external_exams:
        idx = all_exams.index(exam)
        row = [exam]
        for i in range(NUM_COS):
            row.append(round(exam_att_dict[exam], 4) if coexam_mat[i, idx] > 0 else "—")
        rows.append(row)
    avg = ["Avg attainment of CO"]
    for i in range(NUM_COS):
        avg.append("—" if np.isnan(co_ext_att[i]) else round(float(co_ext_att[i]), 4))
    return _tbl(headers, rows, avg)


def render_final_table(co_names, co_int_att, co_ext_att, co_finals, iw, uw, external_label="University"):
    ir, ur = normalize_weights(iw, uw)
    headers = ["Process"] + co_names
    rows = []
    r1 = [f"{uw:.1f}% of {external_label}"]
    r2 = [f"{iw:.1f}% of Internal"]
    for i in range(NUM_COS):
        r1.append("—" if np.isnan(co_ext_att[i]) else round(float(co_ext_att[i]) * ur, 4))
        r2.append("—" if np.isnan(co_int_att[i]) else round(float(co_int_att[i]) * ir, 4))
    rows = [r1, r2]
    fin = ["CO Attainment"] + [round(float(v), 4) for v in co_finals]
    return _tbl(headers, rows, fin)


def render_di_table(exam_di_dict, top_pct, bot_pct, threshold, internal_exams):
    headers = ["Exam", "Type", "N", f"Top {top_pct:.0f}% (H)", f"Bot {bot_pct:.0f}% (L)",
               "k(top)", "k(bot)", DI_LABEL, "Status"]
    rows = []
    for exam, d in exam_di_dict.items():
        t    = "Internal" if exam in internal_exams else "External"
        ok   = d["DI"] >= threshold
        stat = f'<span class="good-di">✓ Good</span>' if ok else f'<span class="warn">⚠ Low</span>'
        rows.append([exam, t, d["N"],
                     f'{d["H"]:.4f}', f'{d["L"]:.4f}',
                     d["k_top"], d["k_bot"],
                     f'<span class="{"good-di" if ok else "warn"}">{d["DI"]:.4f}</span>',
                     stat])
    # Don't pass through _v for pre-formatted HTML cells; build directly
    thead = "".join(f'<th class="proc">{h}</th>' if i == 0 else f"<th>{h}</th>"
                    for i, h in enumerate(headers))
    tbody = ""
    for row in rows:
        cells = "".join(f'<td class="proc">{row[0]}</td>' +
                        "".join(f"<td>{v}</td>" for v in row[1:]))
        tbody += f"<tr>{cells}</tr>"
    return f'<div class="nba-wrap"><table class="nba-tbl"><thead><tr>{thead}</tr></thead><tbody>{tbody}</tbody></table></div>'


def render_po_table(co_names, co_finals, po_contrib, po_finals):
    po_names = [f"PO{j+1}" for j in range(NUM_POS)]
    headers = ["CO", "CO Att."] + po_names
    rows = []
    for i, co in enumerate(co_names):
        att = co_finals[i]
        cls = "hi" if att >= 2.0 else "warn"
        att_cell = f'<span class="{cls}">{att:.4f}</span>'
        cells = [co, att_cell] + [
            f'<span class="zero">0</span>' if po_contrib[i,j] == 0 else f"{po_contrib[i,j]:.4f}"
            for j in range(NUM_POS)
        ]
        rows.append(cells)
    avg = ["Average", ""] + [
        f'<span class="{"hi" if v >= 2.0 else "warn" if v > 0 else "zero"}">{v:.4f}</span>'
        for v in po_finals
    ]
    # Build raw (already HTML-formatted)
    thead = '<th class="proc">CO</th><th>CO Att.</th>' + "".join(f"<th>{p}</th>" for p in po_names)
    tbody = ""
    for row in rows:
        cells = f'<td class="proc">{row[0]}</td>' + "".join(f"<td>{v}</td>" for v in row[1:])
        tbody += f"<tr>{cells}</tr>"
    avg_cells = f'<td class="proc">Average</td>' + "".join(f"<td>{v}</td>" for v in avg[1:])
    tbody += f'<tr class="avg-r">{avg_cells}</tr>'
    return f'<div class="nba-wrap"><table class="nba-tbl"><thead><tr>{thead}</tr></thead><tbody>{tbody}</tbody></table></div>'


def render_pso_table(co_names, co_finals, pso_contrib, pso_finals):
    pso_names = [f"PSO{j+1}" for j in range(NUM_PSOS)]
    headers = ["CO", "CO Att."] + pso_names
    rows = []
    for i, co in enumerate(co_names):
        att = co_finals[i]
        cls = "hi" if att >= 2.0 else "warn"
        att_cell = f'<span class="{cls}">{att:.4f}</span>'
        cells = [co, att_cell] + [
            f'<span class="zero">0</span>' if pso_contrib[i,j] == 0 else f"{pso_contrib[i,j]:.4f}"
            for j in range(NUM_PSOS)
        ]
        rows.append(cells)
    avg = ["Average", ""] + [
        f'<span class="{"hi" if v >= 2.0 else "warn" if v > 0 else "zero"}">{v:.4f}</span>'
        for v in pso_finals
    ]
    thead = '<th class="proc">CO</th><th>CO Att.</th>' + "".join(f"<th>{p}</th>" for p in pso_names)
    tbody = ""
    for row in rows:
        cells = f'<td class="proc">{row[0]}</td>' + "".join(f"<td>{v}</td>" for v in row[1:])
        tbody += f"<tr>{cells}</tr>"
    avg_cells = f'<td class="proc">Average</td>' + "".join(f"<td>{v}</td>" for v in avg[1:])
    tbody += f'<tr class="avg-r">{avg_cells}</tr>'
    return f'<div class="nba-wrap"><table class="nba-tbl"><thead><tr>{thead}</tr></thead><tbody>{tbody}</tbody></table></div>'


# ══════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════

def export_excel(R, subj_name, top_pct, bot_pct, di_threshold) -> BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb  = Workbook()
    ws0 = wb.active

    def hdr(ws, r, c, val, fg=NAV, sz=11, bold=True, wrap=True, center=True):
        cell = ws.cell(r, c, val)
        cell.font      = Font(bold=bold, color="FFFFFF", size=sz, name="Calibri")
        cell.fill      = PatternFill("solid", fgColor=fg)
        cell.alignment = Alignment(
            horizontal="center" if center else "left",
            vertical="center", wrap_text=wrap)
        return cell

    def dat(ws, r, c, val, bold=False, even=False, center=True, color=None):
        cell = ws.cell(r, c, val)
        cell.font      = Font(bold=bold, color=color or "000000", size=10, name="Calibri")
        cell.fill      = PatternFill("solid", fgColor="F7F9FD" if even else "FFFFFF")
        cell.alignment = Alignment(
            horizontal="center" if center else "left", vertical="center")
        return cell

    def avg_cell(ws, r, c, val):
        cell = ws.cell(r, c, val)
        cell.font      = Font(bold=True, color="000000", size=10, name="Calibri")
        cell.fill      = PatternFill("solid", fgColor=LBL)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        return cell

    thin = Side(style="thin", color="D0D7E8")
    bd   = Border(bottom=thin, right=thin, top=thin, left=thin)

    def apply_bd(ws, r1, r2, c1, c2):
        for row in ws.iter_rows(r1, r2, c1, c2):
            for cell in row:
                cell.border = bd

    def set_col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def section_title(ws, r, c1, c2, text):
        ws.merge_cells(f"{get_column_letter(c1)}{r}:{get_column_letter(c2)}{r}")
        cell = ws.cell(r, c1, text)
        cell.font      = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
        cell.fill      = PatternFill("solid", fgColor=NAV)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[r].height = 22
        return r + 1

    def write_df_block(ws, start_r, df, title, avg_label="Average Attainment of CO"):
        """Write a table-block; last row treated as average if it matches avg_label."""
        co_names = R["co_names"]
        ncols = len(df.columns)
        r = section_title(ws, start_r, 1, ncols, title)
        # Headers
        for ci, col in enumerate(df.columns, 1):
            c = hdr(ws, r, ci, col, fg=BLU)
            ws.column_dimensions[get_column_letter(ci)].width = max(
                ws.column_dimensions[get_column_letter(ci)].width or 10,
                min(len(str(col)) + 4, 22))
        r += 1
        for di_row, (_, row_data) in enumerate(df.iterrows()):
            even = di_row % 2 == 0
            is_avg = str(row_data.iloc[0]).lower().startswith(("average", "co attainment"))
            for ci, val in enumerate(row_data, 1):
                if is_avg:
                    c = avg_cell(ws, r, ci, val if val != "—" else "")
                    if ci == 1:
                        c.font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
                        c.fill = PatternFill("solid", fgColor=BLU)
                        c.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    c = dat(ws, r, ci, val if val != "—" else "", even=even,
                            center=(ci != 1), bold=(ci == 1))
            r += 1
        apply_bd(ws, start_r + 1, r - 1, 1, ncols)
        return r + 1

    # ── Sheet 1: Cover ───────────────────────────────────────────────────
    ws0.title = "Cover"
    ws0.sheet_view.showGridLines = False
    ws0.column_dimensions["A"].width = 28
    ws0.column_dimensions["B"].width = 60

    ws0.merge_cells("A1:B1")
    t = ws0.cell(1, 1, "CO-PO Attainment Report  |  NBA Accreditation Format")
    t.font      = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
    t.fill      = PatternFill("solid", fgColor=NAV)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws0.row_dimensions[1].height = 32

    info = [
        ("Subject Code",      R["subj_code"]),
        ("Subject Name",      subj_name),
        ("Assessment Mode",   R.get("mode", "Theory")),
        ("Number of COs",     NUM_COS),
        ("Number of POs",     NUM_POS),
        ("Number of PSOs",    NUM_PSOS),
        ("Total Students",    len(R["df"])),
        ("Internal Weight",   f"{R['iw']:.1f}%"),
        (f"{R.get('external_label', 'University')} Weight", f"{R['uw']:.1f}%"),
        ("DI Top Group",      f"{top_pct:.0f}%"),
        ("DI Bottom Group",   f"{bot_pct:.0f}%"),
        ("DI Threshold",      f"{di_threshold:.2f}"),
    ]
    for ri, (lbl, val) in enumerate(info, 2):
        l = ws0.cell(ri, 1, lbl)
        l.font = Font(bold=True, size=11, name="Calibri")
        l.fill = PatternFill("solid", fgColor=LBL)
        v = ws0.cell(ri, 2, val)
        v.font = Font(size=11, name="Calibri")
        ws0.row_dimensions[ri].height = 18

    r = len(info) + 3
    ws0.cell(r, 1, "CO Statements").font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
    ws0.cell(r, 1).fill = PatternFill("solid", fgColor=BLU)
    ws0.row_dimensions[r].height = 20
    r += 1
    for i, (co, stmt) in enumerate(zip(R["co_names"], R["co_stmts"])):
        ws0.cell(r, 1, co).font = Font(bold=True, size=10, name="Calibri")
        ws0.cell(r, 1).fill = PatternFill("solid", fgColor="EEF3FF" if i % 2 == 0 else "FFFFFF")
        c = ws0.cell(r, 2, stmt)
        c.font      = Font(size=10, name="Calibri")
        c.alignment = Alignment(wrap_text=True)
        ws0.row_dimensions[r].height = 20
        r += 1

    r += 1
    ws0.cell(r, 1, "Maximum Marks per Exam").font = Font(bold=True, size=12, color="FFFFFF", name="Calibri")
    ws0.cell(r, 1).fill = PatternFill("solid", fgColor=BLU)
    r += 1
    for exam, mark in R["max_marks"].items():
        ws0.cell(r, 1, exam).font = Font(bold=True, size=10, name="Calibri")
        ws0.cell(r, 2, mark).font = Font(size=10, name="Calibri")
        r += 1

    # ── Sheet 2: Student Data ────────────────────────────────────────────
    ws2 = wb.create_sheet("Student Data")
    ws2.sheet_view.showGridLines = False
    cols = ["Student Name"] + R["all_exams"]
    for ci, col in enumerate(cols, 1):
        hdr(ws2, 1, ci, col)
        ws2.column_dimensions[get_column_letter(ci)].width = 16
    ws2.row_dimensions[1].height = 20
    for ri, row in R["df"].iterrows():
        even = ri % 2 == 0
        dat(ws2, ri+2, 1, row.get("Student Name", f"Std_{ri+1}"), even=even, center=False)
        for ci, exam in enumerate(R["all_exams"], 2):
            v = row.get(exam, 0)
            dat(ws2, ri+2, ci, round(float(v), 2) if pd.notna(v) else 0, even=even)
    apply_bd(ws2, 1, len(R["df"])+1, 1, len(cols))

    # ── Sheet 3: CO Attainment ───────────────────────────────────────────
    ws3 = wb.create_sheet("CO Attainment")
    ws3.sheet_view.showGridLines = False
    int_df  = build_internal_df(R["co_names"], R["coexam_matrix"],
                                R["exam_att_dict"], R["co_int_att"],
                                R["internal_exams"], R["all_exams"])
    ext_df  = build_external_df(R["co_names"], R["coexam_matrix"],
                                R["exam_att_dict"], R["co_ext_att"],
                                R["external_exams"], R["all_exams"])
    fin_df  = build_final_df(R["co_names"], R["co_int_att"], R["co_ext_att"],
                             R["co_finals"], R["iw"], R["uw"], R.get("external_label", "University"))
    r = write_df_block(ws3, 1, int_df, "a) Attainment of CO through Internal Assessment")
    r = write_df_block(ws3, r, ext_df, f"b) Attainment of CO through {R.get('external_label', 'University')} Assessment")
    write_df_block(ws3, r, fin_df,
                   f"c) Actual CO Attainment = {R['uw']:.1f}% {R.get('external_label', 'University')} + {R['iw']:.1f}% Internal")

    # ── Sheet 4: Discrimination Index ────────────────────────────────────
    ws4 = wb.create_sheet("Discrimination Index")
    ws4.sheet_view.showGridLines = False
    di_df = build_di_df(R["exam_di_dict"], top_pct, bot_pct, di_threshold, R["internal_exams"])
    section_title(ws4, 1, 1, len(di_df.columns),
                  f"Discrimination Index  |  Top {top_pct:.0f}% / Bottom {bot_pct:.0f}%  |  Threshold = {di_threshold:.2f}")
    for ci, col in enumerate(di_df.columns, 1):
        hdr(ws4, 2, ci, col, fg=BLU)
        ws4.column_dimensions[get_column_letter(ci)].width = max(len(str(col)) + 3, 14)
    ws4.row_dimensions[2].height = 20
    for ri, (_, row) in enumerate(di_df.iterrows(), 3):
        even = ri % 2 == 0
        for ci, val in enumerate(row, 1):
            clean_val = str(val).replace("✓ ", "").replace("⚠ ", "")
            c = dat(ws4, ri, ci, clean_val, even=even, center=(ci != 1))
            is_di = (ci == len(di_df.columns) - 1)
            if is_di:
                di_val = float(di_df.iloc[ri-3][DI_LABEL])
                c.font = Font(bold=True, size=10, name="Calibri",
                              color="16A34A" if di_val >= di_threshold else "DC2626")
    apply_bd(ws4, 2, 2+len(di_df), 1, len(di_df.columns))

    # ── Sheet 5: PO Attainment ───────────────────────────────────────────
    ws5 = wb.create_sheet("PO Attainment")
    ws5.sheet_view.showGridLines = False
    po_df = build_po_df(R["co_names"], R["co_finals"],
                        R["po_contrib"], R["po_finals"])
    write_df_block(ws5, 1, po_df, "PO Attainment  [Formula: (CO_att/3) × mapping weight  |  Average over all COs]")

    # ── Sheet 6: PSO Attainment ──────────────────────────────────────────
    ws6 = wb.create_sheet("PSO Attainment")
    ws6.sheet_view.showGridLines = False
    pso_df = build_pso_df(R["co_names"], R["co_finals"],
                          R["pso_contrib"], R["pso_finals"])
    write_df_block(ws6, 1, pso_df, "PSO Attainment  [Formula: (CO_att/3) × mapping weight  |  Average over all COs]")

    # ── Sheet 7: Summary ────────────────────────────────────────────────
    ws7 = wb.create_sheet("Summary")
    ws7.sheet_view.showGridLines = False
    ws7.column_dimensions["A"].width = 20
    for ci in range(2, 7):
        ws7.column_dimensions[get_column_letter(ci)].width = 14

    section_title(ws7, 1, 1, 5, "CO Attainment Summary")
    for ci, h in enumerate(["CO", "Internal", R.get("external_label", "University"), "Final", "Target Met"], 1):
        hdr(ws7, 2, ci, h, fg=BLU)
    for ri, co in enumerate(R["co_names"], 3):
        i = ri - 3
        even = i % 2 == 0
        int_v = R["co_int_att"][i]
        ext_v = R["co_ext_att"][i]
        fin_v = R["co_finals"][i]
        dat(ws7, ri, 1, co, bold=True, even=even, center=False)
        dat(ws7, ri, 2, "—" if np.isnan(int_v) else round(float(int_v), 4), even=even)
        dat(ws7, ri, 3, "—" if np.isnan(ext_v) else round(float(ext_v), 4), even=even)
        dat(ws7, ri, 4, round(float(fin_v), 4), even=even,
            color="1740AD" if fin_v >= 2.0 else "DC2626", bold=True)
        dat(ws7, ri, 5, "Yes" if fin_v >= 2.0 else "No", even=even,
            color="16A34A" if fin_v >= 2.0 else "DC2626", bold=True)
    apply_bd(ws7, 2, 2+NUM_COS, 1, 5)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════
#  PDF EXPORT (A4 PORTRAIT — reportlab)
# ══════════════════════════════════════════════════════════════════════════

def _rl_color(hex_str):
    h = hex_str.lstrip("#")
    return colors.Color(int(h[0:2],16)/255, int(h[2:4],16)/255, int(h[4:6],16)/255)

C_NAV  = _rl_color("061231")
C_BLU  = _rl_color("1740AD")
C_LBL  = _rl_color("EEF3FF")
C_EVN  = _rl_color("F7F9FD")
C_RED  = _rl_color("DC2626")
C_GRN  = _rl_color("16A34A")
C_WHT  = colors.white
C_BLK  = colors.black
C_GREY = _rl_color("D0D7E8")


def _rl_table(data, col_widths, highlight_last=True):
    """Build a formatted reportlab Table from list-of-lists data."""
    n_cols = len(data[0])
    style = [
        ("BACKGROUND",   (0, 0), (-1, 0),    C_NAV),
        ("TEXTCOLOR",    (0, 0), (-1, 0),    C_WHT),
        ("FONTNAME",     (0, 0), (-1, 0),    "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, 0),    8),
        ("ALIGN",        (0, 0), (-1, -1),   "CENTER"),
        ("ALIGN",        (0, 1), (0, -1),    "LEFT"),
        ("VALIGN",       (0, 0), (-1, -1),   "MIDDLE"),
        ("ROWBACKGROUND",(0, 1), (-1, -1),   [C_WHT, C_EVN]),
        ("GRID",         (0, 0), (-1, -1),   0.4, C_GREY),
        ("FONTNAME",     (0, 1), (-1, -1),   "Helvetica"),
        ("FONTSIZE",     (0, 1), (-1, -1),   7.5),
        ("LEFTPADDING",  (0, 0), (-1, -1),   5),
        ("RIGHTPADDING", (0, 0), (-1, -1),   5),
        ("TOPPADDING",   (0, 0), (-1, -1),   4),
        ("BOTTOMPADDING",(0, 0), (-1, -1),   4),
        ("BACKGROUND",   (0, 0), (0, -1),    C_BLU),
        ("TEXTCOLOR",    (0, 1), (0, -1),    C_WHT),
        ("FONTNAME",     (0, 1), (0, -1),    "Helvetica-Bold"),
    ]
    if highlight_last and len(data) > 2:
        style += [
            ("BACKGROUND",  (0, -1), (-1, -1), C_LBL),
            ("FONTNAME",    (0, -1), (-1, -1), "Helvetica-Bold"),
            ("FONTSIZE",    (0, -1), (-1, -1), 7.5),
            ("BACKGROUND",  (0, -1), (0, -1),  C_BLU),
            ("TEXTCOLOR",   (0, -1), (0, -1),  C_WHT),
            ("LINEABOVE",   (0, -1), (-1, -1), 1.5, C_BLU),
        ]
    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle(style))
    return tbl


def _section_para(text, styles):
    style = ParagraphStyle("sec", parent=styles["Normal"],
                           fontSize=11, fontName="Helvetica-Bold",
                           textColor=C_NAV, spaceAfter=4, spaceBefore=10,
                           borderPadding=(4, 0, 4, 10),
                           borderColor=C_BLU, borderWidth=0)
    return Paragraph(f"<b>{text}</b>", style)


def export_pdf(R, subj_name, top_pct, bot_pct, di_threshold) -> BytesIO:
    if not REPORTLAB_OK:
        return _fallback_pdf(R, subj_name)

    buf = BytesIO()
    styles = getSampleStyleSheet()
    W, H = A4   # 595 × 842 points
    margin = 1.8 * cm
    usable_w = W - 2 * margin

    title_style = ParagraphStyle("title", parent=styles["Normal"],
        fontSize=18, fontName="Helvetica-Bold", textColor=C_NAV,
        spaceAfter=6, alignment=TA_LEFT)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"],
        fontSize=10, fontName="Helvetica", textColor=_rl_color("334155"),
        spaceAfter=14, alignment=TA_LEFT)
    body_style = ParagraphStyle("body", parent=styles["Normal"],
        fontSize=9, fontName="Helvetica", textColor=C_BLK,
        spaceAfter=4, leading=14)
    bold_style = ParagraphStyle("boldb", parent=styles["Normal"],
        fontSize=9.5, fontName="Helvetica-Bold", textColor=C_NAV,
        spaceAfter=3, spaceBefore=8)
    small_style = ParagraphStyle("small", parent=styles["Normal"],
        fontSize=8, fontName="Helvetica", textColor=_rl_color("6b7280"),
        spaceAfter=2)

    def page_header_footer(canvas, doc):
        canvas.saveState()
        canvas.setFillColor(C_NAV)
        canvas.rect(margin, H - margin + 2*mm, usable_w, 0.3*mm, fill=1, stroke=0)
        canvas.setFont("Helvetica-Bold", 8)
        canvas.setFillColor(C_NAV)
        canvas.drawString(margin, H - margin + 4*mm,
            f"{R['subj_code']} | {subj_name}  —  CO-PO Attainment Report (NBA Format)")
        canvas.drawRightString(W - margin, H - margin + 4*mm, f"v3.0")
        # Footer
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(_rl_color("94a3b8"))
        canvas.drawString(margin, 14*mm,
            "CO-PO Attainment System | NBA Accreditation Tool")
        canvas.drawRightString(W - margin, 14*mm, f"Page {doc.page}")
        canvas.setFillColor(C_BLU)
        canvas.rect(margin, 12*mm, usable_w, 0.3*mm, fill=1, stroke=0)
        canvas.restoreState()

    story = []

    # ─── COVER PAGE ──────────────────────────────────────────────────────
    story.append(Spacer(1, 1.5*cm))
    story.append(Paragraph("CO-PO Attainment Report", title_style))
    story.append(Paragraph(
        f"{R['subj_code']}  |  {subj_name}  |  COs: {NUM_COS}  |  POs: {NUM_POS}  |  PSOs: {NUM_PSOS}",
        sub_style))
    story.append(HRFlowable(width="100%", thickness=1.5, color=C_BLU))
    story.append(Spacer(1, 0.5*cm))

    meta = [
        ["Total Students", str(len(R["df"])), "Internal Weight", f"{R['iw']:.1f}%"],
        [f"{R.get('external_label', 'University')} Weight", f"{R['uw']:.1f}%", "DI Top Group", f"{top_pct:.0f}%"],
        ["DI Bottom Group", f"{bot_pct:.0f}%", "DI Threshold", f"{di_threshold:.2f}"],
    ]
    meta_tbl = Table(meta, colWidths=[4*cm, 4*cm, 4*cm, 4*cm])
    meta_tbl.setStyle(TableStyle([
        ("GRID",        (0,0),(-1,-1), 0.5, C_GREY),
        ("FONTNAME",    (0,0),(0,-1),  "Helvetica-Bold"),
        ("FONTNAME",    (2,0),(2,-1),  "Helvetica-Bold"),
        ("FONTSIZE",    (0,0),(-1,-1), 9),
        ("ROWBACKGROUND",(0,0),(-1,-1),[C_EVN, C_WHT]),
        ("ALIGN",       (0,0),(-1,-1),"LEFT"),
        ("TOPPADDING",  (0,0),(-1,-1), 5),
        ("BOTTOMPADDING",(0,0),(-1,-1),5),
        ("LEFTPADDING", (0,0),(-1,-1), 8),
    ]))
    story.append(meta_tbl)
    story.append(Spacer(1, 0.8*cm))

    story.append(Paragraph("Course Outcomes", bold_style))
    co_data = [["CO", "Statement"]]
    for co, stmt in zip(R["co_names"], R["co_stmts"]):
        co_data.append([co, stmt])
    co_tbl = Table(co_data, colWidths=[3.5*cm, usable_w - 3.5*cm])
    co_tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0,0),(-1,0),  C_BLU),
        ("TEXTCOLOR",    (0,0),(-1,0),  C_WHT),
        ("FONTNAME",     (0,0),(-1,0),  "Helvetica-Bold"),
        ("FONTNAME",     (0,1),(-1,-1), "Helvetica"),
        ("FONTNAME",     (0,1),(0,-1),  "Helvetica-Bold"),
        ("FONTSIZE",     (0,0),(-1,-1), 9),
        ("ROWBACKGROUND",(0,1),(-1,-1), [C_WHT, C_EVN]),
        ("GRID",         (0,0),(-1,-1), 0.4, C_GREY),
        ("ALIGN",        (0,0),(-1,-1), "LEFT"),
        ("VALIGN",       (0,0),(-1,-1), "MIDDLE"),
        ("TOPPADDING",   (0,0),(-1,-1), 5),
        ("BOTTOMPADDING",(0,0),(-1,-1), 5),
        ("LEFTPADDING",  (0,0),(-1,-1), 6),
    ]))
    story.append(co_tbl)
    story.append(PageBreak())

    # ─── ATTAINMENT LEVEL REFERENCE ──────────────────────────────────────
    story.append(_section_para("Attainment Level Reference", styles))
    lvl_data = [["Level", "Percentage", "Interpretation"],
                ["0", "< 40%", "Not Achieved"],
                ["1", "40–55%", "Partially Achieved"],
                ["2", "56–70%", "Achieved"],
                ["3", "> 70%", "Highly Achieved"]]
    lvl_tbl = Table(lvl_data, colWidths=[3*cm, 4*cm, 6*cm])
    lvl_tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0,0),(-1,0),  C_NAV),
        ("TEXTCOLOR",    (0,0),(-1,0),  C_WHT),
        ("FONTNAME",     (0,0),(-1,0),  "Helvetica-Bold"),
        ("FONTNAME",     (0,1),(-1,-1), "Helvetica"),
        ("FONTSIZE",     (0,0),(-1,-1), 9),
        ("GRID",         (0,0),(-1,-1), 0.4, C_GREY),
        ("ALIGN",        (0,0),(-1,-1), "CENTER"),
        ("TOPPADDING",   (0,0),(-1,-1), 5),
        ("BOTTOMPADDING",(0,0),(-1,-1), 5),
    ]))
    story.append(lvl_tbl)
    story.append(Spacer(1, 0.5*cm))

    # ─── HELPER: df → reportlab table ────────────────────────────────────
    def df_to_rl(df, title, col_widths=None, highlight_last=True):
        story.append(_section_para(title, styles))
        data = [list(df.columns)] + df.fillna("—").astype(str).values.tolist()
        if col_widths is None:
            col_widths = [usable_w / len(df.columns)] * len(df.columns)
            col_widths[0] = min(col_widths[0] * 1.6, 4.5*cm)
        story.append(_rl_table(data, col_widths, highlight_last))
        story.append(Spacer(1, 0.4*cm))

    # ─── CO ATTAINMENT (INTERNAL) ────────────────────────────────────────
    int_df = build_internal_df(R["co_names"], R["coexam_matrix"],
                               R["exam_att_dict"], R["co_int_att"],
                               R["internal_exams"], R["all_exams"])
    cw_int = [3.8*cm] + [(usable_w - 3.8*cm) / NUM_COS] * NUM_COS
    df_to_rl(int_df, "a) Attainment of CO through Internal Assessment", cw_int)

    ext_df = build_external_df(R["co_names"], R["coexam_matrix"],
                               R["exam_att_dict"], R["co_ext_att"],
                               R["external_exams"], R["all_exams"])
    df_to_rl(ext_df, f"b) Attainment of CO through {R.get('external_label', 'University')} Assessment", cw_int)

    fin_df = build_final_df(R["co_names"], R["co_int_att"], R["co_ext_att"],
                            R["co_finals"], R["iw"], R["uw"], R.get("external_label", "University"))
    df_to_rl(fin_df,
             f"c) CO Attainment = {R['uw']:.1f}% {R.get('external_label', 'University')} + {R['iw']:.1f}% Internal",
             cw_int)
    story.append(PageBreak())

    # ─── DISCRIMINATION INDEX ────────────────────────────────────────────
    di_df = build_di_df(R["exam_di_dict"], top_pct, bot_pct, di_threshold, R["internal_exams"])
    # Remove emoji from Status for PDF
    di_df["Status"] = di_df["Status"].str.replace("✓ ", "Good ").str.replace("⚠ ", "Low ")
    di_cw = [3.2*cm, 2.2*cm, 1.5*cm, 2.5*cm, 2.5*cm, 1.5*cm, 1.5*cm, 2.5*cm, 2*cm]
    story.append(_section_para(
        f"Discrimination Index  |  Top {top_pct:.0f}% / Bot {bot_pct:.0f}%  |  Threshold: {di_threshold:.2f}",
        styles))
    di_data = [list(di_df.columns)] + di_df.values.tolist()
    di_tbl = Table(di_data, colWidths=di_cw, repeatRows=1)
    di_style = [
        ("BACKGROUND",  (0,0),(-1,0),  C_NAV),
        ("TEXTCOLOR",   (0,0),(-1,0),  C_WHT),
        ("FONTNAME",    (0,0),(-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",    (0,0),(-1,-1), 7.5),
        ("FONTNAME",    (0,1),(-1,-1), "Helvetica"),
        ("ALIGN",       (0,0),(-1,-1), "CENTER"),
        ("ALIGN",       (0,1),(0,-1),  "LEFT"),
        ("GRID",        (0,0),(-1,-1), 0.4, C_GREY),
        ("TOPPADDING",  (0,0),(-1,-1), 4),
        ("BOTTOMPADDING",(0,0),(-1,-1),4),
        ("LEFTPADDING", (0,0),(-1,-1), 4),
    ]
    for ri, (exam, d) in enumerate(R["exam_di_dict"].items(), 1):
        ok = d["DI"] >= di_threshold
        fc = C_GRN if ok else C_RED
        di_style.append(("TEXTCOLOR", (-2, ri), (-1, ri), fc))
        di_style.append(("FONTNAME",  (-2, ri), (-1, ri), "Helvetica-Bold"))
        if ri % 2 == 0:
            di_style.append(("BACKGROUND", (0, ri), (-1, ri), C_EVN))
    di_tbl.setStyle(TableStyle(di_style))
    story.append(di_tbl)
    story.append(Spacer(1, 0.5*cm))
    story.append(PageBreak())

    # ─── PO ATTAINMENT ───────────────────────────────────────────────────
    po_df = build_po_df(R["co_names"], R["co_finals"],
                        R["po_contrib"], R["po_finals"])
    po_cw = [2.8*cm, 1.8*cm] + [(usable_w - 4.6*cm) / NUM_POS] * NUM_POS
    df_to_rl(po_df, "PO Attainment  [contrib = (CO_att/3) × mapping,  avg over all COs]", po_cw)

    pso_df = build_pso_df(R["co_names"], R["co_finals"],
                          R["pso_contrib"], R["pso_finals"])
    pso_cw = [3*cm, 2*cm] + [(usable_w - 5*cm) / NUM_PSOS] * NUM_PSOS
    df_to_rl(pso_df, "PSO Attainment  [same formula]", pso_cw)
    story.append(PageBreak())

    # --- KEY SUMMARY --------------------------------------------------------
    co_ok = int((R["co_finals"] >= 2.0).sum())
    avg_co = float(np.mean(R["co_finals"])) if len(R["co_finals"]) else 0.0
    po_v = R["po_finals"]
    avg_po = float(po_v[po_v > 0].mean()) if (po_v > 0).any() else 0.0
    low_di = sum(1 for d in R["exam_di_dict"].values() if d["DI"] < di_threshold)
    story.append(_section_para("Key Summary", styles))
    summary_data = [
        ["Metric", "Value"],
        ["Average CO Attainment", f"{avg_co:.3f} / 3.000"],
        ["COs at Target", f"{co_ok} / {NUM_COS} (target >= 2.0)"],
        ["Average PO Attainment", f"{avg_po:.3f} / 3.000"],
        ["Low DI Exams", f"{low_di} / {len(R['all_exams'])} below {di_threshold:.2f}"],
    ]
    story.append(_rl_table(summary_data, [7*cm, usable_w - 7*cm], highlight_last=False))
    story.append(PageBreak())

    # --- CHARTS: one clear A4 section per visual ---------------------------
    def fig_to_img(fig):
        buf2 = BytesIO()
        fig.savefig(buf2, format="png", dpi=170, bbox_inches="tight",
                    facecolor="white")
        plt.close(fig)
        buf2.seek(0)
        return buf2

    def append_chart(title, fig, max_height=17.5*cm):
        width = usable_w
        height = min(width * (fig.get_figheight() / fig.get_figwidth()), max_height)
        story.append(_section_para(title, styles))
        story.append(Image(fig_to_img(fig), width=width, height=height))
        story.append(PageBreak())

    colors_co = [("#1740AD" if v >= 2.5 else "#22c55e" if v >= 2 else
                  "#f59e0b" if v >= 1.5 else "#ef4444") for v in R["co_finals"]]

    fig, ax = plt.subplots(figsize=(7.4, 4.8))
    fig.patch.set_facecolor("white")
    ax.bar(R["co_names"], np.round(R["co_finals"], 3), color=colors_co)
    ax.axhline(2.0, linestyle="--", color="#ef4444", lw=1.3, label="Target 2.0")
    ax.axhline(avg_co, linestyle=":", color="#1740AD", lw=1.5, label=f"CO Avg {avg_co:.3f}")
    ax.set_title("Final CO Attainment", fontsize=13, fontweight="bold")
    ax.set_ylabel("Attainment")
    ax.set_ylim(0, 3.4)
    ax.grid(axis="y", ls=":", alpha=0.35)
    ax.legend(frameon=False, fontsize=9)
    for i, v in enumerate(R["co_finals"]):
        ax.text(i, v + 0.05, f"{v:.3f}", ha="center", fontsize=9)
    plt.tight_layout(pad=1.8)
    append_chart("Chart 1: Final CO Attainment", fig)

    fig, ax = plt.subplots(figsize=(7.4, 4.8))
    fig.patch.set_facecolor("white")
    x = np.arange(NUM_COS)
    w = 0.36
    int_plot = np.where(np.isnan(R["co_int_att"]), 0, R["co_int_att"])
    ext_plot = np.where(np.isnan(R["co_ext_att"]), 0, R["co_ext_att"])
    ax.bar(x - w/2, np.round(int_plot, 3), width=w, color="#5c6bc0",
           label=f"Internal ({R['iw']:.0f}%)")
    ax.bar(x + w/2, np.round(ext_plot, 3), width=w, color="#0891b2",
           label=f"{R.get('external_label', 'University')} ({R['uw']:.0f}%)")
    ax.set_xticks(x)
    ax.set_xticklabels(R["co_names"])
    ax.set_title(f"Internal vs {R.get('external_label', 'University')} Attainment", fontsize=13, fontweight="bold")
    ax.set_ylabel("Attainment")
    ax.set_ylim(0, 3.4)
    ax.grid(axis="y", ls=":", alpha=0.35)
    ax.legend(frameon=False, fontsize=9)
    plt.tight_layout(pad=1.8)
    append_chart(f"Chart 2: Internal vs {R.get('external_label', 'University')} Attainment", fig)

    po_names = [f"PO{j+1}" for j in range(NUM_POS)]
    fig, ax = plt.subplots(figsize=(7.4, 4.8))
    fig.patch.set_facecolor("white")
    ax.bar(po_names, np.round(R["po_finals"], 3),
           color=[("#1740AD" if v >= 2.5 else "#22c55e" if v >= 2 else
                   "#f59e0b" if v >= 1 else "#ef4444") for v in R["po_finals"]])
    ax.axhline(2.0, ls="--", color="#ef4444", lw=1.3, label="Target 2.0")
    ax.set_title("PO Attainment", fontsize=13, fontweight="bold")
    ax.set_ylabel("Attainment")
    ax.set_ylim(0, 3.4)
    ax.tick_params(axis="x", rotation=30)
    ax.grid(axis="y", ls=":", alpha=0.35)
    ax.legend(frameon=False, fontsize=9)
    for i, v in enumerate(R["po_finals"]):
        ax.text(i, v + 0.05, f"{v:.2f}", ha="center", fontsize=8)
    plt.tight_layout(pad=1.8)
    append_chart("Chart 3: PO Attainment", fig)

    pso_names = [f"PSO{j+1}" for j in range(NUM_PSOS)]
    fig, ax = plt.subplots(figsize=(7.4, 4.8))
    fig.patch.set_facecolor("white")
    ax.bar(pso_names, np.round(R["pso_finals"], 3),
           color=["#1740AD", "#0f766e", "#b45309"][:NUM_PSOS])
    ax.axhline(2.0, ls="--", color="#ef4444", lw=1.3, label="Target 2.0")
    ax.set_title("PSO Attainment", fontsize=13, fontweight="bold")
    ax.set_ylabel("Attainment")
    ax.set_ylim(0, 3.4)
    ax.grid(axis="y", ls=":", alpha=0.35)
    ax.legend(frameon=False, fontsize=9)
    for i, v in enumerate(R["pso_finals"]):
        ax.text(i, v + 0.05, f"{v:.3f}", ha="center", fontsize=10)
    plt.tight_layout(pad=1.8)
    append_chart("Chart 4: PSO Attainment", fig)

    fig, ax = plt.subplots(figsize=(7.4, 4.8))
    fig.patch.set_facecolor("white")
    di_vals = [R["exam_di_dict"][e]["DI"] for e in R["all_exams"]]
    di_clrs = ["#22c55e" if d >= di_threshold else "#ef4444" for d in di_vals]
    ax.bar(R["all_exams"], di_vals, color=di_clrs)
    ax.axhline(di_threshold, ls="--", color="#f59e0b", lw=1.5,
               label=f"Threshold ({di_threshold:.2f})")
    ax.set_title("Discrimination Index per Exam", fontsize=13, fontweight="bold")
    ax.set_ylabel("DI")
    ax.set_ylim(0, max(max(di_vals)*1.35, di_threshold*2))
    ax.grid(axis="y", ls=":", alpha=0.35)
    ax.legend(frameon=False, fontsize=9)
    for i, v in enumerate(di_vals):
        ax.text(i, v + 0.002, f"{v:.4f}", ha="center", fontsize=8)
    plt.tight_layout(pad=1.8)
    append_chart("Chart 5: Discrimination Index", fig)

    fig, ax = plt.subplots(figsize=(7.4, 4.8))
    fig.patch.set_facecolor("white")
    heat = np.round(R["po_contrib"], 3)
    im = ax.imshow(heat, cmap="Blues", aspect="auto")
    ax.set_xticks(np.arange(NUM_POS))
    ax.set_xticklabels(po_names, rotation=45, ha="right")
    ax.set_yticks(np.arange(NUM_COS))
    ax.set_yticklabels(R["co_names"])
    ax.set_title("CO to PO Contribution Heatmap", fontsize=13, fontweight="bold")
    for i in range(NUM_COS):
        for j in range(NUM_POS):
            ax.text(j, i, f"{heat[i, j]:.2f}", ha="center", va="center", fontsize=7,
                    color="white" if heat[i, j] > heat.max() * 0.55 else "#061231")
    fig.colorbar(im, ax=ax, fraction=0.035, pad=0.02, label="Contribution")
    plt.tight_layout(pad=1.8)
    append_chart("Chart 6: CO to PO Contribution Heatmap", fig)

    # --- BUILD PDF ---------------------------------------------------------
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=margin, rightMargin=margin,
        topMargin=margin + 0.8*cm, bottomMargin=margin + 0.6*cm,
    )
    doc.build(story, onFirstPage=page_header_footer, onLaterPages=page_header_footer)
    buf.seek(0)
    return buf


def _fallback_pdf(R, subj_name) -> BytesIO:
    """Simple matplotlib PDF if reportlab not available."""
    from matplotlib.backends.backend_pdf import PdfPages
    buf = BytesIO()
    with PdfPages(buf) as pdf:
        fig, ax = plt.subplots(figsize=(8.27, 11.69))
        ax.axis("off")
        ax.text(.5, .95, "CO-PO Attainment Report", ha="center", fontsize=18,
                fontweight="bold", transform=ax.transAxes)
        ax.text(.5, .88, f"{R['subj_code']} | {subj_name}", ha="center",
                fontsize=12, transform=ax.transAxes, color="#334155")
        ax.text(.1, .82, "Install reportlab for a full A4 PDF:\n  pip install reportlab",
                fontsize=11, transform=ax.transAxes, color="#7c3aed")
        pdf.savefig(fig, bbox_inches="tight")
        plt.close(fig)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════
#  PLOTLY CHARTS
# ══════════════════════════════════════════════════════════════════════════

_BASE = dict(
    plot_bgcolor="white", paper_bgcolor="white",
    font=dict(family="Inter, sans-serif", size=12, color="#1a1a2e"),
    margin=dict(l=10, r=10, t=50, b=10), height=360,
)


def _cc(v):
    if v < 1.0: return "#ef4444"
    if v < 1.5: return "#f97316"
    if v < 2.0: return "#f59e0b"
    if v < 2.5: return "#22c55e"
    return "#1740ad"


def chart_co_final(co_names, co_finals):
    avg_co = float(np.mean(co_finals)) if len(co_finals) else 0.0
    fig = go.Figure(go.Bar(
        x=co_names, y=np.round(co_finals, 3),
        marker_color=[_cc(v) for v in co_finals],
        text=np.round(co_finals, 3), textposition="outside",
    ))
    fig.add_hline(y=2.0, line_dash="dash", line_color="#ef4444",
                  annotation_text="Target 2.0", annotation_position="top right")
    fig.add_hline(y=avg_co, line_dash="dot", line_color="#1740ad",
                  annotation_text=f"CO Avg {avg_co:.3f}",
                  annotation_position="bottom right")
    fig.update_layout(**_BASE,
        title=dict(text="<b>Final CO Attainment</b>",
                   font=dict(size=15, color="#061231"), x=0),
        xaxis_title="Course Outcomes",
        yaxis=dict(title="Attainment", range=[0, 3.6]),
    )
    return fig


def chart_int_ext(co_names, co_int, co_ext, iw, uw, external_label="University"):
    int_v = np.where(np.isnan(co_int), 0, np.round(co_int, 3))
    ext_v = np.where(np.isnan(co_ext), 0, np.round(co_ext, 3))
    fig = go.Figure([
        go.Bar(name=f"Internal ({iw:.0f}%)", x=co_names, y=int_v,
               marker_color="#5c6bc0", text=int_v, textposition="outside"),
        go.Bar(name=f"{external_label} ({uw:.0f}%)", x=co_names, y=ext_v,
               marker_color="#0097a7", text=ext_v, textposition="outside"),
    ])
    fig.update_layout(**_BASE,
        title=dict(text=f"<b>Internal vs {external_label} Attainment</b>",
                   font=dict(size=15, color="#061231"), x=0),
        barmode="group", xaxis_title="Course Outcomes",
        yaxis=dict(title="Attainment", range=[0, 3.6]),
        legend=dict(orientation="h", y=1.05, x=0),
    )
    return fig


def chart_po(po_finals):
    po_names = [f"PO{j+1}" for j in range(NUM_POS)]
    fig = go.Figure(go.Bar(
        x=po_names, y=np.round(po_finals, 3),
        marker_color=[_cc(v) for v in po_finals],
        text=np.round(po_finals, 3), textposition="outside",
    ))
    fig.add_hline(y=2.0, line_dash="dash", line_color="#ef4444",
                  annotation_text="Target 2.0", annotation_position="top right")
    fig.update_layout(**_BASE,
        title=dict(text="<b>PO Attainment</b>", font=dict(size=15, color="#061231"), x=0),
        xaxis_title="Program Outcomes",
        yaxis=dict(title="Attainment", range=[0, max(po_finals.max()*1.35, 0.5)]),
    )
    return fig


def chart_pso(pso_finals):
    pso_names = [f"PSO{j+1}" for j in range(NUM_PSOS)]
    fig = go.Figure(go.Bar(
        x=pso_names, y=np.round(pso_finals, 3),
        marker_color=["#1740ad", "#0f766e", "#b45309"][:NUM_PSOS],
        text=np.round(pso_finals, 3), textposition="outside", width=0.45,
    ))
    fig.add_hline(y=2.0, line_dash="dash", line_color="#ef4444",
                  annotation_text="Target 2.0", annotation_position="top right")
    fig.update_layout(**_BASE,
        title=dict(text="<b>PSO Attainment</b>", font=dict(size=15, color="#061231"), x=0),
        xaxis_title="Program Specific Outcomes",
        yaxis=dict(title="Attainment", range=[0, max(pso_finals.max()*1.35, 0.5)]),
    )
    return fig


def chart_radar(co_names, co_int, co_ext, co_finals, external_label="University"):
    int_v = np.where(np.isnan(co_int), 0, co_int)
    ext_v = np.where(np.isnan(co_ext), 0, co_ext)
    cats = co_names + [co_names[0]]
    def w(a): return list(a) + [a[0]]
    fig = go.Figure([
        go.Scatterpolar(r=w(co_finals), theta=cats, name="Final",
                        fill="toself", line_color="#1740ad", fillcolor="rgba(23,64,173,.18)"),
        go.Scatterpolar(r=w(int_v), theta=cats, name="Internal",
                        fill="toself", line_color="#5c6bc0", fillcolor="rgba(92,107,192,.12)"),
        go.Scatterpolar(r=w(ext_v), theta=cats, name=external_label,
                        fill="toself", line_color="#0891b2", fillcolor="rgba(8,145,178,.12)"),
    ])
    fig.update_layout(
        polar=dict(radialaxis=dict(visible=True, range=[0, 3])),
        title=dict(text="<b>CO Attainment Radar</b>",
                   font=dict(size=15, color="#061231"), x=0),
        paper_bgcolor="white", height=380,
        font=dict(family="Inter"),
        legend=dict(orientation="h", y=-0.18, x=0),
        margin=dict(l=40, r=40, t=50, b=50),
    )
    return fig


def chart_di(exam_di_dict, threshold, all_exams):
    di_vals = [exam_di_dict[e]["DI"] for e in all_exams]
    fig = go.Figure(go.Bar(
        x=all_exams, y=di_vals,
        marker_color=["#22c55e" if d >= threshold else "#ef4444" for d in di_vals],
        text=[f"{d:.4f}" for d in di_vals], textposition="outside",
    ))
    fig.add_hline(y=threshold, line_dash="dash", line_color="#f59e0b",
                  annotation_text=f"Threshold ({threshold:.2f})",
                  annotation_position="top right")
    fig.update_layout(**_BASE,
        title=dict(text="<b>Discrimination Index per Exam</b>",
                   font=dict(size=15, color="#061231"), x=0),
        xaxis_title="Exam",
        yaxis=dict(title="DI", range=[0, max(max(di_vals)*1.4, threshold*2)]),
    )
    return fig


def chart_co_heatmap(co_names, co_finals, copo_matrix):
    po_names = [f"PO{j+1}" for j in range(NUM_POS)]
    contrib = np.array([[co_finals[i]/3*copo_matrix[i,j]
                         for j in range(NUM_POS)]
                        for i in range(NUM_COS)])
    fig = go.Figure(go.Heatmap(
        z=np.round(contrib, 3), x=po_names, y=co_names,
        colorscale=[[0,"#f8fafc"],[0.3,"#bfdbfe"],[0.7,"#3b82f6"],[1,"#061231"]],
        text=np.round(contrib, 3), texttemplate="%{text}",
        showscale=True, colorbar=dict(title="Contrib"),
    ))
    fig.update_layout(**{**_BASE, "height": 300},
        title=dict(text="<b>CO→PO Contribution Heatmap</b>",
                   font=dict(size=15, color="#061231"), x=0),
        xaxis_title="Program Outcomes", yaxis_title="Course Outcomes",
    )
    return fig


# ══════════════════════════════════════════════════════════════════════════
#  SAMPLE EXCEL TEMPLATE
# ══════════════════════════════════════════════════════════════════════════

def generate_sample_excel(mode="Theory", components=None) -> BytesIO:
    np.random.seed(99)
    n = 60
    if mode == "Practical":
        components = components or []
        data = {
            "Student Name": [f"Student_{i+1:02d}" for i in range(n)],
            PRACTICAL_INTERNAL_EXAM: np.random.randint(8, 26, n),
        }
        for i, comp in enumerate(components):
            col = practical_component_col(i)
            upper = int(comp.get("max_marks", 20)) + 1
            data[col] = np.random.randint(max(0, upper // 3), max(1, upper), n)
        df = pd.DataFrame(data)
    else:
        df = pd.DataFrame({
            "Student Name": [f"Student_{i+1:02d}" for i in range(n)],
            "Unit Test 1":  np.random.randint(8,  31, n),
            "Unit Test 2":  np.random.randint(10, 31, n),
            "Unit Test 3":  np.random.randint(10, 31, n),
            "Prelim":       np.random.randint(25, 71, n),
            "Insem":        np.random.randint(8,  31, n),
            "Endsem":       np.random.randint(25, 71, n),
        })
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name="Student Marks")
        from openpyxl.styles import Font, PatternFill, Alignment
        ws = w.sheets["Student Marks"]
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF", size=11)
            cell.fill      = PatternFill("solid", fgColor="061231")
            cell.alignment = Alignment(horizontal="center")
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = \
                max(len(str(c.value or "")) for c in col) + 4
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════
#  SESSION STATE
# ══════════════════════════════════════════════════════════════════════════

def init_state():
    defs = {
        "processed":         False,
        "results":           {},
        "df":                None,
        "max_marks":         dict(DEFAULT_MAX),
        "subj_code":         DEFAULT_SUBJ_CODE,
        "subj_name":         DEFAULT_SUBJ_NAME,
        "co_stmts":          list(DEFAULT_CO_STMTS),
        "iw":                DEFAULT_INTERNAL_WEIGHT,
        "uw":                DEFAULT_UNIVERSITY_WEIGHT,
        "di_top_pct":        DEFAULT_DI_TOP_PCT,
        "di_bot_pct":        DEFAULT_DI_BOT_PCT,
        "di_threshold":      DEFAULT_DI_THRESHOLD,
        "mode":              DEFAULT_MODE,
        "practical_internal_max": DEFAULT_PRACTICAL_INTERNAL_MAX,
        "practical_components": [dict(c) for c in DEFAULT_PRACTICAL_COMPONENTS],
    }
    for k, v in defs.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_state()


# ══════════════════════════════════════════════════════════════════════════
#  CALCULATION ORCHESTRATOR
# ══════════════════════════════════════════════════════════════════════════

def run_calculations(df, max_marks, copo_mat, copso_mat, coexam_mat,
                     subj_code, co_stmts, iw, uw,
                     di_top_pct, di_bot_pct, mode="Theory"):
    try:
        co_names = [f"CO{i+1}" for i in range(NUM_COS)]
        ir, ur   = normalize_weights(iw, uw)
        _, ext_label = assessment_labels(mode)

        internal_exams, external_exams, all_exams = get_exam_sets(mode)

        exam_att_dict = {}
        exam_di_dict  = {}
        for exam in all_exams:
            marks = df[exam].fillna(0).values
            att, _ = exam_attainment(marks, max_marks[exam])
            DI, H, L, k_top, k_bot = discrimination_index(marks, max_marks[exam], di_top_pct, di_bot_pct)
            exam_att_dict[exam] = att
            exam_di_dict[exam]  = {"DI": DI, "H": H, "L": L,
                                   "k_top": k_top, "k_bot": k_bot, "N": len(marks)}

        co_int_att, co_ext_att = compute_co_attainments(
            exam_att_dict, coexam_mat, internal_exams, external_exams, all_exams
        )
        co_finals = compute_co_finals(co_int_att, co_ext_att, ir, ur)
        po_contrib,  po_finals  = compute_po_pso(co_finals, copo_mat)
        pso_contrib, pso_finals = compute_po_pso(co_finals, copso_mat)

        st.session_state.results = {
            "co_names":      co_names,
            "co_stmts":      co_stmts,
            "subj_code":     subj_code,
            "exam_att_dict": exam_att_dict,
            "exam_di_dict":  exam_di_dict,
            "co_int_att":    co_int_att,
            "co_ext_att":    co_ext_att,
            "co_finals":     co_finals,
            "po_contrib":    po_contrib,
            "po_finals":     po_finals,
            "pso_contrib":   pso_contrib,
            "pso_finals":    pso_finals,
            "copo_matrix":   copo_mat,
            "copso_matrix":  copso_mat,
            "coexam_matrix": coexam_mat,
            "max_marks":     max_marks,
            "iw": iw, "uw": uw,
            "df": df,
            "mode": mode,
            "external_label": ext_label,
            "internal_exams": internal_exams,
            "external_exams": external_exams,
            "all_exams": all_exams,
        }
        st.session_state.processed = True

    except Exception as e:
        import traceback
        st.error(f"Calculation error: {e}")
        st.code(traceback.format_exc())
        st.session_state.processed = False


# ══════════════════════════════════════════════════════════════════════════
#  MAIN APP
# ══════════════════════════════════════════════════════════════════════════

def run_app():
    iw = st.session_state.iw
    uw = st.session_state.uw
    mode = st.session_state.mode

    # ── HEADER ─────────────────────────────────────────────────────────
    st.markdown(f"""
    <div class="nba-hdr">
      <div class="hdr-icon">🎓</div>
      <div class="hdr-text">
        <div class="hdr-title">CO-PO Attainment &amp; Academic Analytics System</div>
        <div class="hdr-sub">NBA Accreditation Tool · v3.0 · Formula Engine: (CO/3)×w | DI=(H-L)/Max Marks</div>
      </div>
      <div class="hdr-pills">
        <div class="hdr-pill">Mode: {mode}</div>
        <div class="hdr-pill">📊 NBA FORMAT</div>
        <div class="hdr-pill">Internal {iw:.0f}% | {assessment_labels(mode)[1]} {uw:.0f}%</div>
      </div>
    </div>
    """, unsafe_allow_html=True)

    # ── TABS ───────────────────────────────────────────────────────────
    t_in, t_res, t_graph, t_exp = st.tabs([
        "  📥  Input & Config  ",
        "  📊  Results (NBA)   ",
        "  📈  Charts          ",
        "  📁  Export          ",
    ])

    # ╔════════════════════════════════════════╗
    # ║  TAB 1 — INPUT                        ║
    # ╚════════════════════════════════════════╝
    with t_in:
        st.markdown('<div class="stitle">🧭 Assessment Mode</div>', unsafe_allow_html=True)
        mode = st.radio("Mode", ["Theory", "Practical"], horizontal=True, index=0 if st.session_state.mode == "Theory" else 1)
        st.session_state.mode = mode
        _, ext_label = assessment_labels(mode)

        practical_components = st.session_state.practical_components
        if mode == "Practical":
            st.markdown('<div class="stitle">🧪 Practical Configuration</div>', unsafe_allow_html=True)
            cpi, cpa = st.columns([1, 1])
            with cpi:
                st.session_state.practical_internal_max = st.number_input(
                    "Internal Max Marks",
                    min_value=1,
                    max_value=500,
                    value=int(st.session_state.practical_internal_max),
                    key="prac_internal_max_input",
                )
            with cpa:
                st.write("")
                if st.button("➕ Add External Component", key="add_practical_component"):
                    practical_components.append({"name": f"Component {len(practical_components)+1}", "max_marks": 25})

            for idx, comp in enumerate(practical_components):
                cx1, cx2, cx3 = st.columns([2, 1, 0.6])
                with cx1:
                    comp["name"] = st.text_input("Component Name", value=comp.get("name", ""), key=f"pc_name_{idx}")
                with cx2:
                    comp["max_marks"] = st.number_input("Max Marks", min_value=1, max_value=500,
                                                        value=int(comp.get("max_marks", 25)), key=f"pc_mm_{idx}")
                with cx3:
                    st.write("")
                    st.write("")
                    if st.button("🗑️", key=f"pc_del_{idx}") and len(practical_components) > 1:
                        practical_components.pop(idx)
                        st.rerun()
            st.session_state.practical_components = practical_components
            st.markdown('<div class="ainfo">External attainment in Practical mode is computed by combining all configured external components per student.</div>', unsafe_allow_html=True)

        internal_exams, external_exams, all_exams = get_exam_sets(mode)
        col_L, col_R = st.columns([1.05, 1], gap="large")

        # ── LEFT COLUMN ───────────────────────────────────────────────
        with col_L:
            st.markdown('<div class="stitle">🏫 Subject Information</div>', unsafe_allow_html=True)
            sc1, sc2 = st.columns([1, 2])
            with sc1:
                subj_code = st.text_input("Subject Code", value=st.session_state.subj_code, key="inp_code")
            with sc2:
                subj_name = st.text_input("Subject Name", value=st.session_state.subj_name, key="inp_name")
            st.session_state.subj_code = subj_code
            st.session_state.subj_name = subj_name

            st.markdown('<div class="stitle">📝 Course Outcomes</div>', unsafe_allow_html=True)
            co_names_cur = [f"CO{i+1}" for i in range(NUM_COS)]
            co_stmts = []
            for i in range(NUM_COS):
                stmt = st.text_input(
                    label=co_names_cur[i],
                    value=st.session_state.co_stmts[i] if i < len(st.session_state.co_stmts) else "",
                    key=f"co_stmt_{i}",
                )
                co_stmts.append(stmt)
            st.session_state.co_stmts = co_stmts

            st.markdown('<div class="stitle">📤 Upload Student Marks</div>', unsafe_allow_html=True)
            req_cols = ["Student Name"] + (all_exams if mode == "Theory" else [PRACTICAL_INTERNAL_EXAM] + practical_component_columns(practical_components))
            st.markdown(f'<div class="ainfo"><b>Required columns:</b> {", ".join(req_cols)}</div>',
                        unsafe_allow_html=True)
            if mode == "Practical":
                comp_note = ", ".join([f"{practical_component_col(i)} ({c['name']})" for i, c in enumerate(practical_components)])
                st.markdown(f'<div class="ainfo"><b>External components:</b> {comp_note}</div>', unsafe_allow_html=True)

            sample_bytes = generate_sample_excel(mode=mode, components=practical_components).getvalue()
            st.download_button("📥 Download Sample Template", data=sample_bytes,
                               file_name=f"sample_marks_{subj_code}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               key="download_sample_template",
                               on_click="ignore")

            uploaded = st.file_uploader("Upload .xlsx / .xls file", type=["xlsx", "xls"])
            if uploaded:
                try:
                    df_raw = pd.read_excel(uploaded)
                    expected = all_exams if mode == "Theory" else [PRACTICAL_INTERNAL_EXAM] + practical_component_columns(practical_components)
                    missing = [c for c in expected if c not in df_raw.columns]
                    if missing:
                        st.error(f"❌ Missing columns: {', '.join(missing)}")
                    else:
                        st.session_state.df = df_raw
                        st.success(f"✅ {len(df_raw)} students loaded!")
                        with st.expander("Preview (first 10 rows)"):
                            st.dataframe(df_raw.head(10), width="stretch")
                except Exception as e:
                    st.error(f"Error reading file: {e}")

            st.markdown('<div class="stitle">⚙️ Maximum Marks per Exam</div>', unsafe_allow_html=True)
            c1, c2, c3 = st.columns(3)
            mm = {}
            mm_source = all_exams if mode == "Theory" else [PRACTICAL_INTERNAL_EXAM] + practical_component_columns(practical_components)
            for idx, exam in enumerate(mm_source):
                with [c1, c2, c3][idx % 3]:
                    default_val = DEFAULT_MAX.get(exam, 100)
                    if mode == "Practical" and exam == PRACTICAL_INTERNAL_EXAM:
                        default_val = st.session_state.practical_internal_max
                    if mode == "Practical" and exam.startswith("External_Component_"):
                        comp_idx = int(exam.split("_")[-1]) - 1
                        if comp_idx < len(practical_components):
                            default_val = practical_components[comp_idx]["max_marks"]
                    mm[exam] = st.number_input(exam, min_value=1, max_value=500,
                                               value=int(st.session_state.max_marks.get(exam, default_val)),
                                               key=f"mm_{exam}")
            if mode == "Practical":
                mm[PRACTICAL_EXTERNAL_EXAM] = sum(mm[c] for c in practical_component_columns(practical_components))
            st.session_state.max_marks = mm

            # Attainment table
            st.markdown('<div class="stitle">📋 Attainment Level Reference</div>', unsafe_allow_html=True)
            st.markdown("""
            <table class="nba-tbl" style="max-width:420px">
              <tr><th>Level</th><th>Percentage</th><th>Meaning</th></tr>
              <tr><td>0</td><td>&lt; 40%</td><td style="color:#b91c1c;font-weight:700">Not Achieved</td></tr>
              <tr><td>1</td><td>40–55%</td><td style="color:#92400e;font-weight:700">Partially Achieved</td></tr>
              <tr><td>2</td><td>56–70%</td><td style="color:#065f46;font-weight:700">Achieved</td></tr>
              <tr><td>3</td><td>&gt; 70%</td><td style="color:#1e3a8a;font-weight:700">Highly Achieved</td></tr>
            </table>""", unsafe_allow_html=True)

        # ── RIGHT COLUMN ──────────────────────────────────────────────
        with col_R:
            st.markdown('<div class="stitle">⚖️ Assessment Weightages</div>', unsafe_allow_html=True)
            w1, w2 = st.columns(2)
            with w1:
                iw = st.number_input("Internal Weight (%)", 0.0, 100.0,
                                     float(st.session_state.iw), 5.0, key="iw_inp")
            with w2:
                uw = st.number_input(f"{ext_label} Weight (%)", 0.0, 100.0,
                                     float(st.session_state.uw), 5.0, key="uw_inp")
            st.session_state.iw = iw
            st.session_state.uw = uw
            ir, ur = normalize_weights(iw, uw)
            if abs(iw + uw - 100) > 0.01:
                st.markdown(f'<div class="ainfo">Total = {iw+uw:.1f}%. Will be normalized to {ir*100:.1f}% / {ur*100:.1f}%.</div>',
                            unsafe_allow_html=True)

            # ── DI CONFIGURATION ──────────────────────────────────────
            st.markdown('<div class="stitle">📐 Discrimination Index Settings</div>', unsafe_allow_html=True)
            st.markdown("""
            <div class="di-config">
              <b>⚙️ Configure DI parameters for your subject</b><br>
              <span style="font-size:12px;color:#555">
                DI = (Mean of top group − Mean of bottom group) / Max Marks<br>
                Adjust group sizes and threshold as per your institutional norms.
              </span>
            </div>""", unsafe_allow_html=True)
            d1, d2, d3 = st.columns(3)
            with d1:
                di_top = st.number_input("Top Group (%)", 5.0, 50.0,
                                         float(st.session_state.di_top_pct), 1.0,
                                         key="di_top",
                                         help="Top N% students used for H (high group)")
            with d2:
                di_bot = st.number_input("Bottom Group (%)", 5.0, 50.0,
                                         float(st.session_state.di_bot_pct), 1.0,
                                         key="di_bot",
                                         help="Bottom N% students used for L (low group)")
            with d3:
                di_thr = st.number_input("DI Threshold", 0.01, 1.0,
                                         float(st.session_state.di_threshold), 0.01,
                                         format="%.2f",
                                         key="di_thr",
                                         help="DI ≥ threshold = Good; DI < threshold = Needs revision")
            st.session_state.di_top_pct   = di_top
            st.session_state.di_bot_pct   = di_bot
            st.session_state.di_threshold = di_thr

            # ── MAPPINGS ──────────────────────────────────────────────
            po_cols  = [f"PO{j+1}"  for j in range(NUM_POS)]
            pso_cols = [f"PSO{j+1}" for j in range(NUM_PSOS)]

            st.markdown('<div class="stitle">🗺️ CO–Exam Mapping (1 = Assessed, 0 = Not Assessed)</div>',
                        unsafe_allow_html=True)
            if mode == "Theory":
                df_coexam_init = pd.DataFrame(DEFAULT_COEXAM, index=co_names_cur, columns=all_exams)
            else:
                df_coexam_init = pd.DataFrame(np.ones((NUM_COS, len(all_exams))), index=co_names_cur, columns=all_exams)
            edited_coexam  = st.data_editor(df_coexam_init, key=f"coexam_ed_{mode}",
                                            width="stretch", num_rows="fixed",
                                            column_config={e: st.column_config.NumberColumn(e, min_value=0, max_value=1, step=1) for e in all_exams})

            st.markdown('<div class="stitle">🗺️ CO–PO Mapping (0 = None to 3 = High)</div>',
                        unsafe_allow_html=True)
            df_copo_init  = pd.DataFrame(DEFAULT_COPO, index=co_names_cur, columns=po_cols)
            edited_copo   = st.data_editor(df_copo_init, key="copo_ed",
                                           width="stretch", num_rows="fixed",
                                           column_config={c: st.column_config.NumberColumn(c, min_value=0, max_value=3, step=1) for c in po_cols})

            st.markdown('<div class="stitle">🗺️ CO–PSO Mapping (0 = None to 3 = High)</div>',
                        unsafe_allow_html=True)
            df_copso_init = pd.DataFrame(DEFAULT_COPSO, index=co_names_cur, columns=pso_cols)
            edited_copso  = st.data_editor(df_copso_init, key="copso_ed",
                                           width="stretch", num_rows="fixed",
                                           column_config={c: st.column_config.NumberColumn(c, min_value=0, max_value=3, step=1) for c in pso_cols})

            st.markdown(f"""
            <div class="ainfo" style="margin-top:14px">
              <b>📐 Formula Reference:</b><br>
              Exam CO Att = (0·P + 1·Q + 2·R + 3·S) / N<br>
              CO Internal = avg of mapped internal exams per CO<br>
              CO {ext_label} = avg of mapped {ext_label.lower()} exams per CO<br>
              Final CO = {iw:.0f}% × Internal + {uw:.0f}% × {ext_label}<br>
              PO<sub>j</sub> = avg<sub>all COs</sub>[ (CO<sub>i</sub>/3) × w<sub>ij</sub> ]<br>
              DI = (H − L) / Max Marks &nbsp;[top {di_top:.0f}% vs bottom {di_bot:.0f}%]
            </div>
            """, unsafe_allow_html=True)

        # ── CALCULATE ─────────────────────────────────────────────────
        st.markdown('<hr class="sec-div">', unsafe_allow_html=True)
        _, btn_col, _ = st.columns([3, 2, 3])
        with btn_col:
            calc = st.button("🚀  Calculate CO-PO Attainment", width="stretch")

        if calc:
            if st.session_state.df is None:
                st.error("⚠️ Please upload a student marks Excel file first.")
            else:
                # clean mappings
                def clean_df(df, lo, hi):
                    df2 = df.copy()
                    for col in df2.columns:
                        df2[col] = pd.to_numeric(df2[col], errors="coerce").fillna(0).clip(lo, hi)
                    return df2
                coexam_clean = clean_df(edited_coexam, 0, 1)
                copo_clean   = clean_df(edited_copo,   0, 3)
                copso_clean  = clean_df(edited_copso,  0, 3)

                with st.spinner("Computing attainment…"):
                    df_calc = st.session_state.df.copy()
                    max_marks_calc = dict(st.session_state.max_marks)
                    if mode == "Practical":
                        component_cols = practical_component_columns(practical_components)
                        df_calc = build_practical_calc_df(df_calc, component_cols)
                        max_marks_calc = {
                            PRACTICAL_INTERNAL_EXAM: float(st.session_state.max_marks[PRACTICAL_INTERNAL_EXAM]),
                            PRACTICAL_EXTERNAL_EXAM: float(sum(st.session_state.max_marks[c] for c in component_cols)),
                        }
                    run_calculations(
                        df_calc,
                        max_marks_calc,
                        copo_clean.values.astype(float),
                        copso_clean.values.astype(float),
                        coexam_clean.values.astype(float),
                        subj_code, co_stmts,
                        iw, uw, di_top, di_bot, mode,
                    )
                if st.session_state.processed:
                    st.success("✅ Calculation complete! Switch to Results, Charts, or Export tabs.")
                    st.balloons()

    # ╔════════════════════════════════════════╗
    # ║  TAB 2 — RESULTS                      ║
    # ╚════════════════════════════════════════╝
    with t_res:
        if not st.session_state.processed:
            st.markdown('<div class="ainfo">👆 Upload data and click <b>Calculate</b> on the Input tab first.</div>',
                        unsafe_allow_html=True)
        else:
            R    = st.session_state.results
            co_n = R["co_names"]
            di_t = st.session_state.di_threshold
            top_p = st.session_state.di_top_pct
            bot_p = st.session_state.di_bot_pct

            # Subject box
            co_items = "".join(
                f'<div class="co-item"><span class="co-lbl">{n}:</span> {s}</div>'
                for n, s in zip(co_n, R["co_stmts"])
            )
            st.markdown(f"""
            <div class="subj-box">
              <div class="sc">Subject Profile</div>
              <div class="sn">{R["subj_code"]} &nbsp;–&nbsp; {st.session_state.subj_name}</div>
              <div class="co-grid">{co_items}</div>
            </div>""", unsafe_allow_html=True)

            # Metric strip
            co_ok  = int((R["co_finals"] >= 2.0).sum())
            avg_co = float(R["co_finals"].mean())
            po_v   = R["po_finals"]
            avg_po = float(po_v[po_v > 0].mean()) if (po_v > 0).any() else 0.0
            low_di = sum(1 for d in R["exam_di_dict"].values() if d["DI"] < di_t)
            n_stu  = len(R["df"])

            st.markdown(f"""
            <div class="mstrip">
              <div class="mbox" style="--ac:#1740ad">
                <div class="ml">Avg CO Attainment</div>
                <div class="mv">{avg_co:.3f}</div>
                <div class="ms">Out of 3.0</div>
              </div>
              <div class="mbox" style="--ac:#22c55e">
                <div class="ml">COs at Target ≥2.0</div>
                <div class="mv">{co_ok} / {NUM_COS}</div>
                <div class="ms">Threshold met</div>
              </div>
              <div class="mbox" style="--ac:#0891b2">
                <div class="ml">Avg PO Attainment</div>
                <div class="mv">{avg_po:.3f}</div>
                <div class="ms">Mapped POs only</div>
              </div>
              <div class="mbox" style="--ac:{'#22c55e' if low_di == 0 else '#ef4444'}">
                <div class="ml">Low DI Exams</div>
                <div class="mv">{low_di} / {len(R["all_exams"])}</div>
                <div class="ms">DI &lt; {di_t:.2f}</div>
              </div>
              <div class="mbox" style="--ac:#7c3aed">
                <div class="ml">Total Students</div>
                <div class="mv">{n_stu}</div>
                <div class="ms">Analysed</div>
              </div>
            </div>
            """, unsafe_allow_html=True)

            st.markdown('<hr class="sec-div">', unsafe_allow_html=True)

            # ── 3.2.2 CO Attainment ──────────────────────────────────
            st.markdown('<div class="stitle">📊 CO Attainment Overview</div>', unsafe_allow_html=True)

            st.markdown("**a) Attainment of CO through Internal Assessment**")
            st.markdown(render_internal_table(co_n, R["coexam_matrix"],
                                              R["exam_att_dict"], R["co_int_att"],
                                              R["internal_exams"], R["all_exams"]),
                        unsafe_allow_html=True)

            st.markdown(f"<br>**b) Attainment of CO through {R.get('external_label', 'University')} Assessment**",
                        unsafe_allow_html=True)
            st.markdown(render_external_table(co_n, R["coexam_matrix"],
                                              R["exam_att_dict"], R["co_ext_att"],
                                              R["external_exams"], R["all_exams"]),
                        unsafe_allow_html=True)

            st.markdown(f"""
            <div class="ainfo" style="margin:10px 0 6px">
              <b>Actual CO Attainment = {R['uw']:.1f}% {R.get('external_label', 'University')} + {R['iw']:.1f}% Internal</b>
            </div>""", unsafe_allow_html=True)
            st.markdown(render_final_table(co_n, R["co_int_att"], R["co_ext_att"],
                                           R["co_finals"], R["iw"], R["uw"], R.get("external_label", "University")),
                        unsafe_allow_html=True)

            st.markdown('<hr class="sec-div">', unsafe_allow_html=True)

            # ── DI ───────────────────────────────────────────────────
            st.markdown(f'<div class="stitle">📋 Discrimination Index &nbsp;<span style="font-weight:500;font-size:12px;color:#555">Top {top_p:.0f}% / Bottom {bot_p:.0f}%  |  Threshold = {di_t:.2f}</span></div>',
                        unsafe_allow_html=True)
            st.markdown(render_di_table(R["exam_di_dict"], top_p, bot_p, di_t, R["internal_exams"]),
                        unsafe_allow_html=True)

            st.markdown('<hr class="sec-div">', unsafe_allow_html=True)

            # ── 3.3.2 PO / PSO ──────────────────────────────────────
            st.markdown('<div class="stitle">🎯 PO / PSO Attainment Overview</div>',
                        unsafe_allow_html=True)
            st.markdown("**PO Attainment** &nbsp;<span style='font-size:11.5px;color:#555;font-weight:400'>contrib = (CO_att/3) × w | Average over all COs</span>",
                        unsafe_allow_html=True)
            st.markdown(render_po_table(co_n, R["co_finals"],
                                        R["po_contrib"], R["po_finals"]),
                        unsafe_allow_html=True)

            st.markdown("<br>**PSO Attainment**", unsafe_allow_html=True)
            st.markdown(render_pso_table(co_n, R["co_finals"],
                                         R["pso_contrib"], R["pso_finals"]),
                        unsafe_allow_html=True)

            st.markdown('<hr class="sec-div">', unsafe_allow_html=True)

            # ── INSIGHTS ─────────────────────────────────────────────
            st.markdown('<div class="stitle">🔍 Insights &amp; Recommendations</div>',
                        unsafe_allow_html=True)
            any_issue = False

            for co, val in zip(co_n, R["co_finals"]):
                if val < 2.0:
                    any_issue = True
                    st.markdown(f"""
                    <div class="awarn">
                      ⚠️ <b>{co} — Attainment LOW ({val:.4f})</b><br>
                      • Revise teaching strategy for this CO<br>
                      • Conduct remedial sessions / extra tutorials<br>
                      • Align exam questions to Bloom's taxonomy
                    </div>""", unsafe_allow_html=True)

            for exam, d in R["exam_di_dict"].items():
                if d["DI"] < di_t:
                    any_issue = True
                    st.markdown(f"""
                    <div class="awarn">
                      📝 <b>{exam} — Low DI = {d['DI']:.4f} (threshold {di_t:.2f})</b><br>
                      • Question paper does not differentiate well<br>
                      • Rebalance difficulty: Easy 30% / Medium 50% / Hard 20%<br>
                      • Add Higher Order Thinking (HOT) questions
                    </div>""", unsafe_allow_html=True)

            good_cos = [(n, v) for n, v in zip(co_n, R["co_finals"]) if v >= 2.5]
            if good_cos:
                st.markdown(f"""
                <div class="agood">
                  ✅ <b>High-Performing COs:</b> {', '.join(f'{n} ({v:.3f})' for n, v in good_cos)}<br>
                  Excellent outcomes for these COs — maintain current approach.
                </div>""", unsafe_allow_html=True)

            if not any_issue:
                st.markdown("""
                <div class="agood">
                  🏆 <b>All CO targets (≥ 2.0) met and DI values healthy!</b><br>
                  Course is on track for NBA accreditation.
                </div>""", unsafe_allow_html=True)

            st.markdown(f"""
            <div class="ainfo">
              <b>📌 Formula Notes:</b> &nbsp;
              Exam CO Att = (0·P + 1·Q + 2·R + 3·S) / N &nbsp;|&nbsp;
              PO/PSO = avg<sub>all COs</sub>[(CO/3) × w] &nbsp;|&nbsp;
              DI = (H − L) / Max Marks with top {top_p:.0f}% / bottom {bot_p:.0f}% groups
            </div>""", unsafe_allow_html=True)

    # ╔════════════════════════════════════════╗
    # ║  TAB 3 — CHARTS                       ║
    # ╚════════════════════════════════════════╝
    with t_graph:
        if not st.session_state.processed:
            st.markdown('<div class="ainfo">👆 Run calculation from the Input tab to view charts.</div>', unsafe_allow_html=True)
        else:
            R    = st.session_state.results
            co_n = R["co_names"]
            di_t = st.session_state.di_threshold

            g1, g2 = st.columns(2, gap="large")
            with g1:
                st.plotly_chart(chart_co_final(co_n, R["co_finals"]),
                                width="stretch")
            with g2:
                st.plotly_chart(chart_int_ext(co_n, R["co_int_att"],
                                              R["co_ext_att"], R["iw"], R["uw"], R.get("external_label", "University")),
                                width="stretch")

            g3, g4 = st.columns(2, gap="large")
            with g3:
                st.plotly_chart(chart_po(R["po_finals"]), width="stretch")
            with g4:
                st.plotly_chart(chart_pso(R["pso_finals"]), width="stretch")

            g5, g6 = st.columns(2, gap="large")
            with g5:
                st.plotly_chart(chart_radar(co_n, R["co_int_att"],
                                            R["co_ext_att"], R["co_finals"], R.get("external_label", "University")),
                                width="stretch")
            with g6:
                st.plotly_chart(chart_di(R["exam_di_dict"], di_t, R["all_exams"]),
                                width="stretch")

            st.plotly_chart(chart_co_heatmap(co_n, R["co_finals"], R["copo_matrix"]),
                            width="stretch")

    # ╔════════════════════════════════════════╗
    # ║  TAB 4 — EXPORT                       ║
    # ╚════════════════════════════════════════╝
    with t_exp:
        if not st.session_state.processed:
            st.markdown('<div class="ainfo">👆 Run calculation from the Input tab to enable report export.</div>', unsafe_allow_html=True)
        else:
            R    = st.session_state.results
            top_p = st.session_state.di_top_pct
            bot_p = st.session_state.di_bot_pct
            di_t  = st.session_state.di_threshold

            st.markdown('<div class="stitle">📁 Download Reports</div>', unsafe_allow_html=True)
            st.markdown(f"""
            <div class="ainfo">
              <b>Excel Report (7 Sheets):</b> Cover · Student Data · CO Attainment (a/b/c) · DI · PO · PSO · Summary<br>
              <b>PDF Report (A4 Portrait):</b> Cover · CO Attainment tables · DI table · PO/PSO tables · Charts<br>
              <span style="font-size:12px;color:#4b5563">Mode-aware labels and values are applied automatically to both exports.</span>
            </div>""", unsafe_allow_html=True)

            with st.spinner("Building Excel + PDF reports…"):
                excel_buf = export_excel(R, st.session_state.subj_name, top_p, bot_p, di_t)
                pdf_buf   = export_pdf(R, st.session_state.subj_name, top_p, bot_p, di_t)
                excel_bytes = excel_buf.getvalue()
                pdf_bytes   = pdf_buf.getvalue()

            d1, d2 = st.columns(2, gap="large")
            with d1:
                st.download_button(
                    "📥  Download Excel Report (.xlsx)",
                    data=excel_bytes,
                    file_name=f"CO_PO_Report_{R['subj_code']}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_excel_report",
                    on_click="ignore",
                    width="stretch",
                )
            with d2:
                st.download_button(
                    "📄  Download PDF Report (A4)",
                    data=pdf_bytes,
                    file_name=f"CO_PO_Report_{R['subj_code']}.pdf",
                    mime="application/pdf",
                    key="download_pdf_report",
                    on_click="ignore",
                    width="stretch",
                )

            st.markdown('<hr class="sec-div">', unsafe_allow_html=True)
            st.markdown('<div class="stitle">📊 Quick Summary Preview</div>', unsafe_allow_html=True)

            p1, p2, p3 = st.columns(3, gap="large")
            with p1:
                st.markdown("**CO Attainment Summary**")
                co_df = pd.DataFrame({
                    "CO":          R["co_names"],
                    "Internal":    ["—" if np.isnan(v) else f"{v:.4f}" for v in R["co_int_att"]],
                    R.get("external_label", "University"):  ["—" if np.isnan(v) else f"{v:.4f}" for v in R["co_ext_att"]],
                    "Final":       np.round(R["co_finals"], 4),
                    "Target":      ["✅ Yes" if v >= 2.0 else "❌ No" for v in R["co_finals"]],
                })
                st.dataframe(co_df, width="stretch", hide_index=True)
            with p2:
                st.markdown("**PO Attainment**")
                po_df = pd.DataFrame({
                    "PO":          [f"PO{j+1}" for j in range(NUM_POS)],
                    "Attainment":  np.round(R["po_finals"], 4),
                    "Status":      ["✅" if v >= 2.0 else "⚠️" if v > 0 else "—"
                                    for v in R["po_finals"]],
                })
                st.dataframe(po_df, width="stretch", hide_index=True)
            with p3:
                st.markdown("**PSO Attainment**")
                pso_df = pd.DataFrame({
                    "PSO":         [f"PSO{j+1}" for j in range(NUM_PSOS)],
                    "Attainment":  np.round(R["pso_finals"], 4),
                    "Status":      ["✅" if v >= 2.0 else "⚠️" if v > 0 else "—"
                                    for v in R["pso_finals"]],
                })
                st.dataframe(pso_df, width="stretch", hide_index=True)


# ══════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    run_app()
